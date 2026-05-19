import datetime as dt
import math
import re
import subprocess
import sys
from copy import copy
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font


INPUT_DIR = Path('input')
OUTPUT_DIR = Path('output')
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

SCRIPT_NAME = '月中_ROI参数文件更新'

# 目标文件固定结构（按当前模板）
TARGET_SHEET_MAIN = '参数更新模板'
TARGET_SHEET_IMPORT = 'Sheet1'
TARGET_SHEET_COMPARE = '参数前后对比'

FORECAST_SHEET_NAME = '参数汇总表 -海外'
BI_SHEET_NAME = 'Sheet1'

# 参数更新模板映射区
MAIN_ROW_START = 2
MAIN_MONTH_COL_START = 6   # F
MAIN_MONTH_COL_END = 17    # Q
MAIN_KEY_COL_START = 19    # S
MAIN_KEY_COL_END = 22      # V

# Sheet1 更新时间
IMPORT_ROW_START = 4
IMPORT_DATE_COL = 20       # T

# 参数前后对比
COMPARE_B4_ROW = 4
COMPARE_B4_COL = 2
COMPARE_HEADER_ROW = 6
COMPARE_DATA_ROW_START = 7
COMPARE_START_COL = 2      # B
COMPARE_END_COL = 10       # J

# BI 文件结构
BI_HEADER_ROW = 4
BI_MONTH_COL = 2
BI_SUBJECT_COL = 3
BI_CHANNEL_COL = 4
BI_METRIC_COL_START = 5    # E
BI_METRIC_COL_END = 10     # J
BI_CREATE_TIME_COL = 21    # U

USE_FILL_SAMPLE = False
DEBUG_LOG_ENABLED = False
ASK_FILE_INPUT = True
RESET_BOLD_BEFORE_RUN = True
ENABLE_BI_EXPORT_STEP = True
BI_EXPORT_SCRIPT_NAME = 'ROI_BI现有参数自动导出.py'


def normalize_text(val) -> str:
    if val is None:
        return ''
    return str(val).strip()

def values_equivalent(a, b, float_tol: float = 1e-9) -> bool:
    if a in (None, '') and b in (None, ''):
        return True

    if isinstance(a, dt.datetime):
        a = a.date()
    if isinstance(b, dt.datetime):
        b = b.date()
    if isinstance(a, dt.date) and isinstance(b, dt.date):
        return a == b

    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        if isinstance(a, float) and isinstance(b, float):
            if math.isnan(a) and math.isnan(b):
                return True
        return abs(float(a) - float(b)) <= float_tol

    return str(a).strip() == str(b).strip()



def is_effective_value(val) -> bool:
    if val in (None, ''):
        return False
    if isinstance(val, str):
        s = val.strip().upper()
        if s.startswith('='):
            return False
        if s in ('#N/A', '#VALUE!', '#REF!', '#DIV/0!', 'N/A'):
            return False
    return True


def get_prefer_computed_value(ws_formula, ws_value, r: int, c: int):
    """
    读取预估源值时优先取 data_only 的计算结果；
    若无缓存结果再回退到公式簿原值。
    """
    v_value = ws_value.cell(r, c).value
    if is_effective_value(v_value):
        return v_value
    formula_cell = ws_formula.cell(r, c)
    # 如果是公式单元格但没有缓存结果，不回退公式本体，避免写入 "=OFFSET(...)" 这类文本
    if formula_cell.data_type == 'f':
        return None
    return formula_cell.value


def normalize_month(val) -> Optional[int]:
    if val is None or val == '':
        return None

    if isinstance(val, dt.datetime):
        return val.year * 100 + val.month
    if isinstance(val, dt.date):
        return val.year * 100 + val.month

    if isinstance(val, (int, float)):
        v = int(val)
        if 190001 <= v <= 299912:
            return v

    s = str(val).strip()
    if not s:
        return None

    m = re.search(r'(20\d{2})\D?(\d{1,2})', s)
    if m:
        y = int(m.group(1))
        mo = int(m.group(2))
        if 1 <= mo <= 12:
            return y * 100 + mo

    m2 = re.search(r'(\d{2})\D?(\d{1,2})', s)
    if m2:
        y2 = int(m2.group(1))
        mo = int(m2.group(2))
        if 1 <= mo <= 12:
            return (2000 + y2) * 100 + mo

    digits = re.sub(r'\D', '', s)
    if len(digits) == 6:
        y = int(digits[:4])
        mo = int(digits[4:])
        if 1 <= mo <= 12:
            return y * 100 + mo

    return None


def debug_write(lines: List[str], out_path: Path):
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text('\n'.join(lines), encoding='utf-8')


def detect_special_col_name(ws_formula, ws_value, col_idx: int) -> str:
    """
    识别预估表第2行的“X月预估数1/2”列名。
    兼容三种情况：
    1) 单元格直接是文本（如 4月预估数2）
    2) 单元格是公式（如 =BP1-1&"月预估数2"）
    3) 兜底固定列：BO=4月预估数2，BP=5月预估数1（当前财务模板实际如此）
    """
    # 优先读 data_only 的展示文本（若文件已计算缓存，最可靠）
    txt = normalize_text(ws_value.cell(2, col_idx).value)
    if txt.endswith('月预估数1') or txt.endswith('月预估数2'):
        return txt

    # 再读公式工作簿中的文本
    raw = ws_formula.cell(2, col_idx).value
    txt_formula = normalize_text(raw)
    if txt_formula.endswith('月预估数1') or txt_formula.endswith('月预估数2'):
        return txt_formula

    # 解析公式中的“xx月预估数1/2”文本片段
    if isinstance(raw, str) and '月预估数' in raw:
        m = re.search(r'月预估数[12]', raw)
        suffix = m.group(0) if m else ''
        if suffix:
            # 从公式中取被引用的基准月份单元格，例如 BP1
            ref = re.search(r'([A-Z]{1,3})1', raw)
            if ref:
                # 基准值优先取 data_only，拿不到再取公式簿
                ref_val = ws_value[f'{ref.group(1)}1'].value
                if ref_val in (None, ''):
                    ref_val = ws_formula[f'{ref.group(1)}1'].value
                try:
                    base_month = int(ref_val)
                    if '-1' in raw:
                        return f'{base_month - 1}{suffix}'
                    return f'{base_month}{suffix}'
                except Exception:
                    pass
    return ''


def fill_key(cell) -> Tuple[object, object, object, object, object]:
    f = cell.fill
    fg = f.fgColor
    return (
        f.patternType,
        fg.type,
        fg.rgb,
        fg.indexed,
        fg.theme,
    )


def learn_target_fill_keys(ws) -> set:
    # 从“应更新候选区”学习底色：B=思维 且 S-V 有值的行的 F-Q
    counts = {}
    for r in range(MAIN_ROW_START, ws.max_row + 1):
        if normalize_text(ws.cell(r, 2).value) != '思维':
            continue
        key_vals = [normalize_text(ws.cell(r, c).value) for c in range(MAIN_KEY_COL_START, MAIN_KEY_COL_END + 1)]
        if not any(key_vals):
            continue
        for c in range(MAIN_MONTH_COL_START, MAIN_MONTH_COL_END + 1):
            k = fill_key(ws.cell(r, c))
            counts[k] = counts.get(k, 0) + 1

    if not counts:
        return set()

    # 取出现最多的前2类样式作为可更新区域
    top = sorted(counts.items(), key=lambda x: x[1], reverse=True)[:2]
    return {k for k, _ in top}


def bold_cell(cell):
    old_font = cell.font
    if old_font is None:
        cell.font = Font(bold=True)
        return
    cell.font = copy(old_font)
    cell.font = cell.font.copy(bold=True)


def unbold_cell(cell):
    old_font = cell.font
    if old_font is None:
        return
    cell.font = copy(old_font)
    cell.font = cell.font.copy(bold=False)


def reset_previous_bold_marks(ws_main, ws_import, ws_compare):
    # 1) 参数更新模板：仅重置脚本会操作的区域
    for r in range(MAIN_ROW_START, ws_main.max_row + 1):
        if normalize_text(ws_main.cell(r, 2).value) != '思维':
            continue
        key = [normalize_text(ws_main.cell(r, c).value) for c in range(MAIN_KEY_COL_START, MAIN_KEY_COL_END + 1)]
        if not any(key):
            continue
        for c in range(MAIN_MONTH_COL_START, MAIN_MONTH_COL_END + 1):
            unbold_cell(ws_main.cell(r, c))

    # 2) Sheet1：T列
    for r in range(IMPORT_ROW_START, ws_import.max_row + 1):
        a = ws_import.cell(r, 1).value
        b = ws_import.cell(r, 2).value
        c = ws_import.cell(r, 3).value
        if a in (None, '') and b in (None, '') and c in (None, ''):
            continue
        unbold_cell(ws_import.cell(r, IMPORT_DATE_COL))

    # 3) 参数前后对比：B4 + B:J数据区
    unbold_cell(ws_compare.cell(COMPARE_B4_ROW, COMPARE_B4_COL))
    for r in range(COMPARE_DATA_ROW_START, ws_compare.max_row + 1):
        for c in range(COMPARE_START_COL, COMPARE_END_COL + 1):
            unbold_cell(ws_compare.cell(r, c))


def choose_file_by_input(
    input_dir: Path,
    prompt: str,
    fallback_path: Optional[Path],
    keyword: Optional[str] = None,
) -> Path:
    candidates = [f for f in input_dir.glob('*.xlsx') if not f.name.startswith('~$')]
    if keyword:
        matched = [f for f in candidates if keyword in f.name]
    else:
        matched = candidates

    # 若关键词未命中，回退到全量列表，避免无文件可选
    show_list = matched if matched else candidates

    print(f'\n{prompt}')
    if keyword:
        print(f'关键词匹配："{keyword}"，命中 {len(matched)} 个')
    for i, f in enumerate(show_list, start=1):
        print(f'  {i}. {f.name}')
    if fallback_path:
        print(f'直接回车使用自动识别: {fallback_path.name}')
    user = input('请输入序号或文件名(回车默认): ').strip()
    if not user:
        if fallback_path:
            return fallback_path
        raise FileNotFoundError('未选择文件且无默认文件')

    if user.isdigit():
        idx = int(user)
        if 1 <= idx <= len(show_list):
            return show_list[idx - 1]
        raise ValueError('输入序号超出范围')

    for f in show_list:
        if f.name == user:
            return f
    # 允许输入关键字
    matches = [f for f in show_list if user in f.name]
    if len(matches) == 1:
        return matches[0]
    if len(matches) > 1:
        raise ValueError('匹配到多个文件，请输入更完整名称或序号')
    raise FileNotFoundError(f'未找到文件: {user}')


def find_files(input_dir: Path, ask_user: bool = False) -> Tuple[Path, Path, Path]:
    xlsx_files = [f for f in input_dir.glob('*.xlsx') if not f.name.startswith('~$')]
    if not xlsx_files:
        raise FileNotFoundError(f'未在 {input_dir} 找到 xlsx 文件')

    forecast_matches = [f for f in xlsx_files if 'ROI计算器固化回写表' in f.name and '预估' in f.name]
    bi_matches = [f for f in xlsx_files if 'BI导出' in f.name and '海外投放ROI系数计算回写' in f.name]
    template_matches = [f for f in xlsx_files if '月中-海外投放ROI系数计算回写模板' in f.name]

    forecast = forecast_matches[0] if forecast_matches else None
    bi_file = bi_matches[0] if bi_matches else None
    template = template_matches[0] if template_matches else None

    if not forecast or not bi_file or not template:
        raise FileNotFoundError(
            '输入文件识别失败，请确认 input 下包含：\n'
            '1) ROI计算器固化回写表-xx月预估-发出.xlsx\n'
            '2) BI导出-海外投放ROI系数计算回写.xlsx\n'
            '3) 月中-海外投放ROI系数计算回写模板-xxxx更新.xlsx'
        )

    # 用户主动要求选择，或某类命中多个候选时，强制选择具体版本
    need_choose = ask_user or len(forecast_matches) > 1 or len(bi_matches) > 1 or len(template_matches) > 1
    if need_choose:
        forecast = choose_file_by_input(
            input_dir,
            '请选择【预估数文件】',
            forecast,
            keyword='ROI计算器固化回写表',
        )
        bi_file = choose_file_by_input(
            input_dir,
            '请选择【BI现有参数文件】',
            bi_file,
            keyword='BI导出-海外投放ROI系数计算回写',
        )
        template = choose_file_by_input(
            input_dir,
            '请选择【目标模板文件】',
            template,
            keyword='月中-海外投放ROI系数计算回写模板',
        )

    return forecast, bi_file, template


def load_forecast_index(forecast_path: Path) -> Dict[Tuple[str, str, str, str], Dict[str, Dict]]:
    wb_formula = load_workbook(forecast_path, data_only=False)
    ws_formula = wb_formula[FORECAST_SHEET_NAME]
    wb_value = load_workbook(forecast_path, data_only=True)
    ws_value = wb_value[FORECAST_SHEET_NAME]
    merged_map = build_merged_value_map(ws_formula)

    month_cols: List[Tuple[int, int]] = []
    # 第二行从 BC 开始
    for c in range(55, ws_formula.max_column + 1):
        m = normalize_month(ws_formula.cell(2, c).value)
        if not m:
            m = normalize_month(ws_value.cell(2, c).value)
        if m:
            month_cols.append((c, m))

    # 识别“X月预估数1/2”专用列（例如 4月预估数2、5月预估数1）
    special_cols = {}
    for c in range(1, ws_formula.max_column + 1):
        head = detect_special_col_name(ws_formula, ws_value, c)
        if head:
            special_cols[head] = c

    idx: Dict[Tuple[str, str, str, str], Dict[str, Dict]] = {}

    for r in range(3, ws_formula.max_row + 1):
        k = (
            normalize_text(get_cell_value(ws_formula, merged_map, r, 3)),
            normalize_text(get_cell_value(ws_formula, merged_map, r, 4)),
            normalize_text(get_cell_value(ws_formula, merged_map, r, 5)),
            normalize_text(get_cell_value(ws_formula, merged_map, r, 6)),
        )
        if not any(k):
            continue

        month_map: Dict[int, object] = {}
        for c, m in month_cols:
            val = get_prefer_computed_value(ws_formula, ws_value, r, c)
            if is_effective_value(val):
                month_map[m] = val

        special_map = {}
        for name, c in special_cols.items():
            val = get_prefer_computed_value(ws_formula, ws_value, r, c)
            if is_effective_value(val):
                special_map[name] = val

        idx[k] = {
            'month_map': month_map,
            'special_map': special_map,
        }

    return idx


def pick_value_with_fallback(
    month_map: Dict[int, object],
    special_map: Dict[str, object],
    month: int,
) -> object:
    if month in month_map and is_effective_value(month_map[month]):
        return month_map[month], f'直取{month}'
    # 新规则：某月份缺值时，优先找“该月份预估数2”，其次“该月份预估数1”
    target_m = month % 100
    name1 = f'{target_m}月预估数1'
    name2 = f'{target_m}月预估数2'

    if name2 in special_map and is_effective_value(special_map[name2]):
        return special_map[name2], f'专用列{name2}'
    if name1 in special_map and is_effective_value(special_map[name1]):
        return special_map[name1], f'专用列{name1}'

    return None, '无可用值'


def update_main_sheet(
    ws,
    forecast_index: Dict[Tuple[str, str, str, str], Dict[str, Dict]],
    run_month: int,
) -> Tuple[int, List[str], Dict[int, Dict[str, int]]]:
    updated = 0
    logs: List[str] = []
    month_stats: Dict[int, Dict[str, int]] = {}
    fill_keys = learn_target_fill_keys(ws) if USE_FILL_SAMPLE else set()
    use_fill_filter = len(fill_keys) > 0
    logs.append(
        f'USE_FILL_SAMPLE={USE_FILL_SAMPLE}, use_fill_filter={use_fill_filter}, '
        f'fill_key_count={len(fill_keys)}, run_month={run_month}'
    )

    month_col_to_month = {}
    for c in range(MAIN_MONTH_COL_START, MAIN_MONTH_COL_END + 1):
        m = normalize_month(ws.cell(1, c).value)
        if m:
            month_col_to_month[c] = m
    logs.append(f'模板月份映射(F-Q): {month_col_to_month}')

    for r in range(MAIN_ROW_START, ws.max_row + 1):
        subject = normalize_text(ws.cell(r, 2).value)
        if subject != '思维':
            continue

        key = tuple(normalize_text(ws.cell(r, c).value) for c in range(MAIN_KEY_COL_START, MAIN_KEY_COL_END + 1))
        if not any(key):
            continue

        if key not in forecast_index:
            logs.append(f'R{r} 键未命中: {key}')
            continue

        row_data = forecast_index[key]
        month_map = row_data.get('month_map', {})
        special_map = row_data.get('special_map', {})
        logs.append(f'R{r} 键命中: {key}, 源月份={sorted(month_map.keys())}, 专用列={list(special_map.keys())}')

        for c, month in month_col_to_month.items():
            if month not in month_stats:
                month_stats[month] = {'updated': 0, 'unchanged': 0, 'no_value': 0}
            cell = ws.cell(r, c)
            # 当前策略：默认不按fill过滤，凡是 B=思维 且 S-V有值 的行，F-Q 都参与更新
            # 若未来需要再开启，可把 USE_FILL_SAMPLE 设为 True
            if use_fill_filter and fill_key(cell) not in fill_keys:
                logs.append(f'R{r}C{c} 月份{month} 跳过: fill不在样本')
                continue

            new_val, source_desc = pick_value_with_fallback(month_map, special_map, month)
            if new_val is None:
                old_val = cell.value
                if old_val not in (None, ''):
                    cell.value = None
                    bold_cell(cell)
                    updated += 1
                    logs.append(f'R{r}C{c} 月份{month} 清空: {old_val} -> 空 ({source_desc})')
                else:
                    logs.append(f'R{r}C{c} 月份{month} 跳过: {source_desc}')
                month_stats[month]['no_value'] += 1
                continue

            if cell.value != new_val:
                logs.append(f'R{r}C{c} 月份{month} 更新: {cell.value} -> {new_val} ({source_desc})')
                cell.value = new_val
                bold_cell(cell)
                updated += 1
                month_stats[month]['updated'] += 1
            else:
                logs.append(f'R{r}C{c} 月份{month} 不变: {new_val} ({source_desc})')
                month_stats[month]['unchanged'] += 1

    return updated, logs, month_stats


def update_sheet1_t_col(ws) -> int:
    updated = 0
    today = dt.datetime.now().date()

    for r in range(IMPORT_ROW_START, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value

        if a in (None, '') and b in (None, '') and c in (None, ''):
            continue

        cell = ws.cell(r, IMPORT_DATE_COL)
        if isinstance(cell.value, dt.datetime):
            old_date = cell.value.date()
        elif isinstance(cell.value, dt.date):
            old_date = cell.value
        else:
            old_date = None

        if old_date != today:
            cell.value = today
            cell.number_format = 'yyyy/mm/dd'
            bold_cell(cell)
            updated += 1

    return updated


def build_merged_value_map(ws):
    merged_map = {}
    for rng in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = rng.bounds
        top_left_val = ws.cell(min_row, min_col).value
        for rr in range(min_row, max_row + 1):
            for cc in range(min_col, max_col + 1):
                merged_map[(rr, cc)] = top_left_val
    return merged_map


def get_cell_value(ws, merged_map, r, c):
    cell = ws.cell(r, c)
    if isinstance(cell, MergedCell):
        return merged_map.get((r, c))
    v = cell.value
    if v in (None, ''):
        return merged_map.get((r, c), v)
    return v


def load_bi_index(bi_path: Path):
    wb = load_workbook(bi_path, data_only=True)
    ws = wb[BI_SHEET_NAME]
    merged_map = build_merged_value_map(ws)

    idx = {}
    latest_create_time = None

    for r in range(BI_HEADER_ROW + 1, ws.max_row + 1):
        month_v = get_cell_value(ws, merged_map, r, BI_MONTH_COL)
        subject_v = get_cell_value(ws, merged_map, r, BI_SUBJECT_COL)
        channel_v = get_cell_value(ws, merged_map, r, BI_CHANNEL_COL)

        month = normalize_month(month_v)
        subject = normalize_text(subject_v)
        channel = normalize_text(channel_v)

        if not month or not subject or not channel:
            continue

        metrics = [get_cell_value(ws, merged_map, r, c) for c in range(BI_METRIC_COL_START, BI_METRIC_COL_END + 1)]
        idx[(month, subject, channel)] = metrics

        ct = get_cell_value(ws, merged_map, r, BI_CREATE_TIME_COL)
        if isinstance(ct, dt.datetime):
            if latest_create_time is None or ct > latest_create_time:
                latest_create_time = ct

    return idx, latest_create_time


def update_compare_sheet(compare_ws, import_ws, bi_idx, latest_time) -> int:
    updated = 0

    if latest_time is not None:
        b4 = compare_ws.cell(COMPARE_B4_ROW, COMPARE_B4_COL)
        if not values_equivalent(b4.value, latest_time):
            b4.value = latest_time
            b4.number_format = 'yyyy/mm/dd'
            bold_cell(b4)
            updated += 1

    # 先清空旧数据，避免残留旧月份（如 202603）
    for rr in range(COMPARE_DATA_ROW_START, compare_ws.max_row + 1):
        for cc in range(COMPARE_START_COL, COMPARE_END_COL + 1):
            cell = compare_ws.cell(rr, cc)
            if cell.value not in (None, ''):
                cell.value = None
                bold_cell(cell)
                updated += 1

    out_row = COMPARE_DATA_ROW_START

    for r in range(IMPORT_ROW_START, import_ws.max_row + 1):
        month = import_ws.cell(r, 1).value
        subject = import_ws.cell(r, 2).value
        channel = import_ws.cell(r, 3).value

        if month in (None, '') and subject in (None, '') and channel in (None, ''):
            continue

        n_month = normalize_month(month)
        n_subject = normalize_text(subject)
        n_channel = normalize_text(channel)

        c_b = compare_ws.cell(out_row, 2)
        c_c = compare_ws.cell(out_row, 3)
        c_d = compare_ws.cell(out_row, 4)

        if not values_equivalent(c_b.value, n_month):
            c_b.value = n_month
            bold_cell(c_b)
            updated += 1
        if not values_equivalent(c_c.value, n_subject):
            c_c.value = n_subject
            bold_cell(c_c)
            updated += 1
        if not values_equivalent(c_d.value, n_channel):
            c_d.value = n_channel
            bold_cell(c_d)
            updated += 1

        metrics = bi_idx.get((n_month, n_subject, n_channel))
        if metrics:
            for i, val in enumerate(metrics, start=5):
                cell = compare_ws.cell(out_row, i)
                if not values_equivalent(cell.value, val):
                    cell.value = val
                    bold_cell(cell)
                    updated += 1

        out_row += 1

    return updated


def main():
    print(f'[{SCRIPT_NAME}] 开始执行...')

    # 步骤0：可选先执行 BI 在线导出
    if ENABLE_BI_EXPORT_STEP:
        run_export = input('是否先自动执行 BI 在线导出？(Y/n，默认Y): ').strip().lower()
        if run_export in ('', 'y', 'yes'):
            export_script = Path(BI_EXPORT_SCRIPT_NAME)
            if not export_script.exists():
                raise FileNotFoundError(f'未找到导出脚本: {export_script.resolve()}')
            print('[STEP 0] 执行 BI 在线导出...')
            subprocess.run([sys.executable, str(export_script)], check=True)
            print('[STEP 0] BI 在线导出完成。')

    ask_mode = ASK_FILE_INPUT
    if ask_mode:
        ans = input('是否手动选择输入文件？(Y/n，默认Y): ').strip().lower()
        ask_mode = (ans in ('', 'y', 'yes'))

    forecast_file, bi_file, template_file = find_files(INPUT_DIR, ask_user=ask_mode)

    debug_enabled = DEBUG_LOG_ENABLED
    ans_debug = input('是否输出调试日志？(y/N，默认N): ').strip().lower()
    if ans_debug in ('y', 'yes'):
        debug_enabled = True
    print(f'[INFO] 预估数文件: {forecast_file.name}')
    print(f'[INFO] BI参数文件: {bi_file.name}')
    print(f'[INFO] 目标模板文件: {template_file.name}')

    forecast_index = load_forecast_index(forecast_file)

    target_wb = load_workbook(template_file, data_only=False)
    ws_main = target_wb[TARGET_SHEET_MAIN]
    ws_import = target_wb[TARGET_SHEET_IMPORT]
    ws_compare = target_wb[TARGET_SHEET_COMPARE]

    reset_bold = RESET_BOLD_BEFORE_RUN
    ans_reset = input('是否先重置历史加粗，仅保留本次变更加粗？(Y/n，默认Y): ').strip().lower()
    if ans_reset in ('n', 'no'):
        reset_bold = False
    if reset_bold:
        reset_previous_bold_marks(ws_main, ws_import, ws_compare)

    today = dt.datetime.now()
    run_month = today.year * 100 + today.month
    cnt_main, main_logs, month_stats = update_main_sheet(ws_main, forecast_index, run_month)
    cnt_sheet1 = update_sheet1_t_col(ws_import)

    bi_idx, latest_time = load_bi_index(bi_file)
    cnt_compare = update_compare_sheet(ws_compare, ws_import, bi_idx, latest_time)

    today_str = today.strftime('%Y%m%d')
    output_name = f'月中-海外投放ROI系数计算回写模板-{today_str}更新.xlsx'
    out_path = OUTPUT_DIR / output_name
    if out_path.exists():
        i = 1
        while True:
            candidate = OUTPUT_DIR / f'月中-海外投放ROI系数计算回写模板-{today_str}更新_v{i}.xlsx'
            if not candidate.exists():
                out_path = candidate
                break
            i += 1
    target_wb.save(out_path)
    debug_log_path = OUTPUT_DIR / f'月中_ROI参数文件更新_debug_{today_str}.log'
    if debug_enabled:
        debug_write(main_logs, debug_log_path)

    print('[SUCCESS] 执行完成')
    print(f'[INFO] 输出文件: {out_path}')
    print(f'[INFO] 参数更新模板更新单元格数: {cnt_main}')
    print(f'[INFO] Sheet1 T列更新单元格数: {cnt_sheet1}')
    print(f'[INFO] 参数前后对比更新单元格数: {cnt_compare}')
    print('[INFO] 月份汇总(更新/不变/无值)：')
    for m in sorted(month_stats.keys()):
        if 202601 <= m <= run_month:
            s = month_stats[m]
            print(f'  {m}: 更新{s["updated"]} / 不变{s["unchanged"]} / 无值{s["no_value"]}')
    if debug_enabled:
        print(f'[INFO] 调试日志: {debug_log_path}')


if __name__ == '__main__':
    main()

