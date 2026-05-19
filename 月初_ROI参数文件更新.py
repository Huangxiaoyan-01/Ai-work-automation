import datetime as dt
import math
import os
import re
import subprocess
import sys
from copy import copy
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font

INPUT_DIR = Path('input')
OUTPUT_DIR = Path('output')
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

SCRIPT_NAME = '月初_ROI参数文件更新'

TARGET_SHEET_MAIN = '参数更新模板'
TARGET_SHEET_IMPORT = 'Sheet1'
TARGET_SHEET_COMPARE = '参数前后对比'

SPLIT_SHEET_NAME = '参数表'
BI_SHEET_NAME = 'Sheet1'

MAIN_ROW_START = 2
MAIN_MONTH_COL_START = 6  # F
MAIN_MONTH_COL_END = 17   # Q
MAIN_KEY_COL_START = 19   # S
MAIN_KEY_COL_END = 22     # V

IMPORT_ROW_START = 4
IMPORT_DATE_COL = 20      # T

COMPARE_B4_ROW = 4
COMPARE_B4_COL = 2
COMPARE_DATA_ROW_START = 7
COMPARE_START_COL = 2
COMPARE_END_COL = 10

BI_HEADER_ROW = 4
BI_MONTH_COL = 2
BI_SUBJECT_COL = 3
BI_CHANNEL_COL = 4
BI_METRIC_COL_START = 5
BI_METRIC_COL_END = 10
BI_CREATE_TIME_COL = 21

ASK_FILE_INPUT = True
RESET_BOLD_BEFORE_RUN = True
ENABLE_BI_EXPORT_STEP = True
BI_EXPORT_SCRIPT_NAME = 'ROI_BI现有参数自动导出.py'


def normalize_text(v) -> str:
    if v is None:
        return ''
    return str(v).strip()


def normalize_month(v) -> Optional[int]:
    if v is None or v == '':
        return None
    if isinstance(v, dt.datetime):
        return v.year * 100 + v.month
    if isinstance(v, dt.date):
        return v.year * 100 + v.month
    if isinstance(v, (int, float)):
        i = int(v)
        if 190001 <= i <= 299912:
            return i
    s = str(v).strip()
    m = re.search(r'(20\d{2})\D?(\d{1,2})', s)
    if m:
        y, mo = int(m.group(1)), int(m.group(2))
        if 1 <= mo <= 12:
            return y * 100 + mo
    m2 = re.search(r'(\d{2})\D?(\d{1,2})', s)
    if m2:
        y2, mo = int(m2.group(1)), int(m2.group(2))
        if 1 <= mo <= 12:
            return (2000 + y2) * 100 + mo
    return None


def extract_month_num_from_text(v) -> Optional[int]:
    s = normalize_text(v)
    if not s:
        return None
    m = re.search(r'(\d{1,2})\s*月', s)
    if m:
        mm = int(m.group(1))
        if 1 <= mm <= 12:
            return mm
    return None


def is_effective_value(v) -> bool:
    if v in (None, ''):
        return False
    if isinstance(v, str):
        s = v.strip().upper()
        if s.startswith('='):
            return False
        if s in ('#N/A', '#VALUE!', '#REF!', '#DIV/0!', 'N/A'):
            return False
    return True


def values_equivalent(a, b, tol: float = 1e-9) -> bool:
    if a in (None, '') and b in (None, ''):
        return True
    if isinstance(a, dt.datetime):
        a = a.date()
    if isinstance(b, dt.datetime):
        b = b.date()
    if isinstance(a, dt.date) and isinstance(b, dt.date):
        return a == b
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        if isinstance(a, float) and isinstance(b, float) and math.isnan(a) and math.isnan(b):
            return True
        return abs(float(a) - float(b)) <= tol
    return str(a).strip() == str(b).strip()


def build_merged_value_map(ws):
    mp = {}
    for rg in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = rg.bounds
        v = ws.cell(min_row, min_col).value
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                mp[(r, c)] = v
    return mp


def get_cell_value(ws, mp, r, c):
    cell = ws.cell(r, c)
    if isinstance(cell, MergedCell):
        return mp.get((r, c))
    v = cell.value
    if v in (None, ''):
        return mp.get((r, c), v)
    return v


def bold_cell(cell):
    old = cell.font
    if old is None:
        cell.font = Font(bold=True)
        return
    cell.font = copy(old)
    cell.font = cell.font.copy(bold=True)


def unbold_cell(cell):
    old = cell.font
    if old is None:
        return
    cell.font = copy(old)
    cell.font = cell.font.copy(bold=False)


def choose_file_by_input(input_dir: Path, prompt: str, fallback: Optional[Path], keyword: Optional[str] = None) -> Path:
    candidates = [f for f in input_dir.glob('*.xlsx') if not f.name.startswith('~$')]
    matched = [f for f in candidates if (keyword in f.name if keyword else True)]
    show = matched if matched else candidates

    print(f'\n{prompt}')
    if keyword:
        print(f'关键词匹配："{keyword}"，命中 {len(matched)} 个')
    for i, f in enumerate(show, start=1):
        print(f'  {i}. {f.name}')

    if fallback:
        print(f'直接回车使用自动识别: {fallback.name}')

    user = input('请输入序号或文件名(回车默认): ').strip()
    if not user:
        if fallback:
            return fallback
        raise FileNotFoundError('未选择文件且无默认文件')

    if user.isdigit():
        idx = int(user)
        if 1 <= idx <= len(show):
            return show[idx - 1]
        raise ValueError('输入序号超出范围')

    for f in show:
        if f.name == user:
            return f

    fuzzy = [f for f in show if user in f.name]
    if len(fuzzy) == 1:
        return fuzzy[0]
    if len(fuzzy) > 1:
        raise ValueError('匹配到多个文件，请输入更完整名称或序号')
    raise FileNotFoundError(f'未找到文件: {user}')


def find_files(input_dir: Path, ask_user: bool = False) -> Tuple[Path, Path, Path]:
    files = [f for f in input_dir.glob('*.xlsx') if not f.name.startswith('~$')]
    if not files:
        raise FileNotFoundError(f'未在 {input_dir} 找到 xlsx 文件')

    split_matches = [f for f in files if '海外目标' in f.name and '思维' in f.name]
    bi_matches = [f for f in files if 'BI导出-海外投放ROI系数计算回写' in f.name]
    template_matches = [f for f in files if '月初-海外投放ROI系数计算回写模板' in f.name]

    split_file = split_matches[0] if split_matches else None
    bi_file = bi_matches[0] if bi_matches else None
    tpl_file = template_matches[0] if template_matches else None

    if not split_file or not bi_file or not tpl_file:
        raise FileNotFoundError('输入文件识别失败，请检查 input 目录文件')

    need_choose = ask_user or len(split_matches) > 1 or len(bi_matches) > 1 or len(template_matches) > 1
    if need_choose:
        split_file = choose_file_by_input(input_dir, '请选择【拆标文件】', split_file, keyword='海外目标')
        bi_file = choose_file_by_input(input_dir, '请选择【BI现有参数文件】', bi_file, keyword='BI导出-海外投放ROI系数计算回写')
        tpl_file = choose_file_by_input(input_dir, '请选择【目标模板文件】', tpl_file, keyword='月初-海外投放ROI系数计算回写模板')

    return split_file, bi_file, tpl_file


def load_split_index(split_path: Path):
    wb_formula = load_workbook(split_path, data_only=False)
    ws = wb_formula[SPLIT_SHEET_NAME]
    wb_value = load_workbook(split_path, data_only=True)
    ws_val = wb_value[SPLIT_SHEET_NAME]

    # F/G 对应月份识别：
    # 1) 优先从标题文本提“X月”
    # 2) 若是公式，尝试解析引用单元格（如 总览!B3 / B8）取日期月份
    # 最终只保留“月号”，后续按目标模板首行的 yyyymm 自动映射
    def detect_month_num(col: int) -> Optional[int]:
        # data_only 可能已算出“4月预估-2”
        mm = extract_month_num_from_text(ws_val.cell(3, col).value)
        if mm:
            return mm
        mm = extract_month_num_from_text(ws.cell(3, col).value)
        if mm:
            return mm

        # 解析公式引用，例如 =LEFT(总览!B3,...)
        formula = normalize_text(ws.cell(3, col).value)
        ref = re.search(r'([\\u4e00-\\u9fa5A-Za-z0-9_]+)!([A-Z]+\\d+)', formula)
        if ref:
            sheet_name, addr = ref.group(1), ref.group(2)
            if sheet_name in wb_val.sheetnames:
                rv = wb_val[sheet_name][addr].value
                if isinstance(rv, (dt.datetime, dt.date)):
                    return rv.month
                mm2 = extract_month_num_from_text(rv)
                if mm2:
                    return mm2
        return None

    m_f_num = detect_month_num(6)
    m_g_num = detect_month_num(7)

    idx = {}
    for r in range(4, ws.max_row + 1):
        # 目标 S-V 需精确匹配拆标 B-E（区域/渠道/科目/项目）
        key = (
            normalize_text(ws.cell(r, 2).value),  # B 区域
            normalize_text(ws.cell(r, 3).value),  # C 渠道
            normalize_text(ws.cell(r, 4).value),  # D 科目
            normalize_text(ws.cell(r, 5).value),  # E 项目
        )
        if not any(key):
            continue

        month_map = {}
        v_f = ws_val.cell(r, 6).value
        if not is_effective_value(v_f):
            v_f = ws.cell(r, 6).value
        if m_f_num and is_effective_value(v_f):
            month_map[m_f_num] = v_f

        v_g = ws_val.cell(r, 7).value
        if not is_effective_value(v_g):
            v_g = ws.cell(r, 7).value
        if m_g_num and is_effective_value(v_g):
            month_map[m_g_num] = v_g

        idx[key] = month_map

    return idx


def update_main_sheet(ws, split_idx):
    updated = 0
    month_stats = {}

    month_col_to_month = {}
    for c in range(MAIN_MONTH_COL_START, MAIN_MONTH_COL_END + 1):
        m = normalize_month(ws.cell(1, c).value)
        if m:
            month_col_to_month[c] = m

    for r in range(MAIN_ROW_START, ws.max_row + 1):
        if normalize_text(ws.cell(r, 2).value) != '思维':
            continue

        key = tuple(normalize_text(ws.cell(r, c).value) for c in range(MAIN_KEY_COL_START, MAIN_KEY_COL_END + 1))
        if not any(key):
            continue

        if key not in split_idx:
            continue

        m_map = split_idx[key]  # key: month_num(1-12) -> value
        for c, m in month_col_to_month.items():
            month_stats.setdefault(m, {'updated': 0, 'unchanged': 0, 'no_value': 0})
            mm = m % 100
            if mm not in m_map:
                month_stats[m]['no_value'] += 1
                continue

            new_val = m_map[mm]
            cell = ws.cell(r, c)
            if not values_equivalent(cell.value, new_val):
                cell.value = new_val
                bold_cell(cell)
                updated += 1
                month_stats[m]['updated'] += 1
            else:
                month_stats[m]['unchanged'] += 1

    return updated, month_stats


def update_sheet1_t_col(ws) -> int:
    updated = 0
    today = dt.datetime.now().date()
    for r in range(IMPORT_ROW_START, ws.max_row + 1):
        a, b, c = ws.cell(r, 1).value, ws.cell(r, 2).value, ws.cell(r, 3).value
        if a in (None, '') and b in (None, '') and c in (None, ''):
            continue
        cell = ws.cell(r, IMPORT_DATE_COL)
        old_date = cell.value.date() if isinstance(cell.value, dt.datetime) else cell.value if isinstance(cell.value, dt.date) else None
        if old_date != today:
            cell.value = today
            cell.number_format = 'yyyy/mm/dd'
            bold_cell(cell)
            updated += 1
    return updated


def load_bi_index(bi_path: Path):
    wb = load_workbook(bi_path, data_only=True)
    ws = wb[BI_SHEET_NAME]
    mp = build_merged_value_map(ws)

    idx = {}
    latest = None
    for r in range(BI_HEADER_ROW + 1, ws.max_row + 1):
        m = normalize_month(get_cell_value(ws, mp, r, BI_MONTH_COL))
        s = normalize_text(get_cell_value(ws, mp, r, BI_SUBJECT_COL))
        ch = normalize_text(get_cell_value(ws, mp, r, BI_CHANNEL_COL))
        if not m or not s or not ch:
            continue

        vals = [get_cell_value(ws, mp, r, c) for c in range(BI_METRIC_COL_START, BI_METRIC_COL_END + 1)]
        idx[(m, s, ch)] = vals

        ct = get_cell_value(ws, mp, r, BI_CREATE_TIME_COL)
        if isinstance(ct, dt.datetime):
            if latest is None or ct > latest:
                latest = ct

    return idx, latest


def update_compare_sheet(compare_ws, import_ws, bi_idx, latest_time):
    updated = 0

    if latest_time is not None:
        b4 = compare_ws.cell(COMPARE_B4_ROW, COMPARE_B4_COL)
        if not values_equivalent(b4.value, latest_time):
            b4.value = latest_time
            b4.number_format = 'yyyy/mm/dd'
            bold_cell(b4)
            updated += 1

    # 先清空旧数据，避免残留旧月份（如 202603）
    for rr in range(COMPARE_DATA_ROW_START, compare_ws.max_row + 1):
        for cc in range(COMPARE_START_COL, COMPARE_END_COL + 1):
            cell = compare_ws.cell(rr, cc)
            if cell.value not in (None, ''):
                cell.value = None
                bold_cell(cell)
                updated += 1

    out_row = COMPARE_DATA_ROW_START
    for r in range(IMPORT_ROW_START, import_ws.max_row + 1):
        month = import_ws.cell(r, 1).value
        subject = import_ws.cell(r, 2).value
        channel = import_ws.cell(r, 3).value
        if month in (None, '') and subject in (None, '') and channel in (None, ''):
            continue

        n_m = normalize_month(month)
        n_s = normalize_text(subject)
        n_c = normalize_text(channel)

        c_b = compare_ws.cell(out_row, 2)
        c_c = compare_ws.cell(out_row, 3)
        c_d = compare_ws.cell(out_row, 4)

        if not values_equivalent(c_b.value, n_m):
            c_b.value = n_m
            bold_cell(c_b)
            updated += 1
        if not values_equivalent(c_c.value, n_s):
            c_c.value = n_s
            bold_cell(c_c)
            updated += 1
        if not values_equivalent(c_d.value, n_c):
            c_d.value = n_c
            bold_cell(c_d)
            updated += 1

        metrics = bi_idx.get((n_m, n_s, n_c))
        if metrics:
            for i, v in enumerate(metrics, start=5):
                cell = compare_ws.cell(out_row, i)
                if not values_equivalent(cell.value, v):
                    cell.value = v
                    bold_cell(cell)
                    updated += 1

        out_row += 1

    return updated


def reset_previous_bold_marks(ws_main, ws_import, ws_compare):
    for r in range(MAIN_ROW_START, ws_main.max_row + 1):
        if normalize_text(ws_main.cell(r, 2).value) != '思维':
            continue
        key = [normalize_text(ws_main.cell(r, c).value) for c in range(MAIN_KEY_COL_START, MAIN_KEY_COL_END + 1)]
        if not any(key):
            continue
        for c in range(MAIN_MONTH_COL_START, MAIN_MONTH_COL_END + 1):
            unbold_cell(ws_main.cell(r, c))

    for r in range(IMPORT_ROW_START, ws_import.max_row + 1):
        a, b, c = ws_import.cell(r, 1).value, ws_import.cell(r, 2).value, ws_import.cell(r, 3).value
        if a in (None, '') and b in (None, '') and c in (None, ''):
            continue
        unbold_cell(ws_import.cell(r, IMPORT_DATE_COL))

    unbold_cell(ws_compare.cell(COMPARE_B4_ROW, COMPARE_B4_COL))
    for r in range(COMPARE_DATA_ROW_START, ws_compare.max_row + 1):
        for c in range(COMPARE_START_COL, COMPARE_END_COL + 1):
            unbold_cell(ws_compare.cell(r, c))


def make_output_path() -> Path:
    today = dt.datetime.now().strftime('%Y%m%d')
    base = OUTPUT_DIR / f'月初-海外投放ROI系数计算回写模板-{today}更新.xlsx'
    if not base.exists():
        return base
    i = 1
    while True:
        p = OUTPUT_DIR / f'月初-海外投放ROI系数计算回写模板-{today}更新_v{i}.xlsx'
        if not p.exists():
            return p
        i += 1


def main():
    print(f'[{SCRIPT_NAME}] 开始执行...')

    if ENABLE_BI_EXPORT_STEP:
        run_export = input('是否先自动执行 BI 在线导出？(Y/n，默认Y): ').strip().lower()
        if run_export in ('', 'y', 'yes'):
            export_script = Path(BI_EXPORT_SCRIPT_NAME)
            if not export_script.exists():
                raise FileNotFoundError(f'未找到导出脚本: {export_script.resolve()}')
            print('[STEP 0] 执行 BI 在线导出...')
            subprocess.run([sys.executable, str(export_script)], check=True)
            print('[STEP 0] BI 在线导出完成。')

    ask_mode = ASK_FILE_INPUT
    if ask_mode:
        ans = input('是否手动选择输入文件？(Y/n，默认Y): ').strip().lower()
        ask_mode = (ans in ('', 'y', 'yes'))

    split_file, bi_file, tpl_file = find_files(INPUT_DIR, ask_user=ask_mode)
    print(f'[INFO] 拆标文件: {split_file.name}')
    print(f'[INFO] BI参数文件: {bi_file.name}')
    print(f'[INFO] 目标模板文件: {tpl_file.name}')

    split_idx = load_split_index(split_file)

    wb = load_workbook(tpl_file, data_only=False)
    ws_main = wb[TARGET_SHEET_MAIN]
    ws_import = wb[TARGET_SHEET_IMPORT]
    ws_compare = wb[TARGET_SHEET_COMPARE]

    reset_bold = RESET_BOLD_BEFORE_RUN
    ans_reset = input('是否先重置历史加粗，仅保留本次变更加粗？(Y/n，默认Y): ').strip().lower()
    if ans_reset in ('n', 'no'):
        reset_bold = False
    if reset_bold:
        reset_previous_bold_marks(ws_main, ws_import, ws_compare)

    cnt_main, month_stats = update_main_sheet(ws_main, split_idx)
    cnt_sheet1 = update_sheet1_t_col(ws_import)

    bi_idx, latest_time = load_bi_index(bi_file)
    cnt_compare = update_compare_sheet(ws_compare, ws_import, bi_idx, latest_time)

    out_path = make_output_path()
    wb.save(out_path)

    print('[SUCCESS] 执行完成')
    print(f'[INFO] 输出文件: {out_path}')
    print(f'[INFO] 参数更新模板更新单元格数: {cnt_main}')
    print(f'[INFO] Sheet1 T列更新单元格数: {cnt_sheet1}')
    print(f'[INFO] 参数前后对比更新单元格数: {cnt_compare}')
    print('[INFO] 月份汇总(更新/不变/无值)：')
    for m in sorted(month_stats.keys()):
        s = month_stats[m]
        print(f'  {m}: 更新{s["updated"]} / 不变{s["unchanged"]} / 无值{s["no_value"]}')


if __name__ == '__main__':
    main()
