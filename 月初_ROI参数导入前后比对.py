import os
import re
import time
import asyncio
import datetime
from pathlib import Path
from dotenv import load_dotenv
from playwright.async_api import async_playwright
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side

BASE_DIR = Path(__file__).resolve().parent

# 加载环境变量
load_dotenv(BASE_DIR / ".env")

# 配置路径
DOWNLOAD_DIR = Path(os.getenv('DOWNLOAD_DIR', 'downloads'))
OUTPUT_DIR = Path(os.getenv('OUTPUT_DIR', 'output'))
DOWNLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

# 报表配置
REPORT_CONFIGS = [
    {
        'name': '投放报表',
        'url': 'https://bi.61info.cn/smartbi/vision/openresource.jsp?resid=I2c928087019462b562b5ce74019482c57b777944',
        'default': True,
        'default_regions': [0, 1, 2, 3, 4]
    },
    {
        'name': '非港澳商务',
        'url': 'https://bi.61info.cn/smartbi/vision/openresource.jsp?resid=I2c928087019462b562b5ce740194833b107e36ec',
        'default': True,
        'default_regions': [0]
    },
    {
        'name': '港澳商务',
        'url': 'https://bi.61info.cn/smartbi/vision/openresource.jsp?resid=I2c928087019462b562b5ce74019484756286034f',
        'default': True,
        'default_regions': [0]
    },
    {
        'name': '台湾商务',
        'url': 'https://bi.61info.cn/smartbi/vision/openresource.jsp?resid=I2c928087019b236723675f9c019b2591839a32ce',
        'default': True,
        'default_regions': [0]
    },
    {
        'name': 'Local商务',
        'url': 'https://bi.61info.cn/smartbi/vision/openresource.jsp?resid=I2c928087019c1fbe1fbefe98019c2819ccf36e2b',
        'default': True,
        'default_regions': [0]
    }
]

# 区域配置
REGION_CONFIGS = [
    {
        'name': '整体含台',
        'regions': ['北美', '欧洲', '海外其他', '亚洲', '澳洲', '港澳', '台湾']
    },
    {
        'name': '整体非台',
        'regions': ['北美', '欧洲', '海外其他', '亚洲', '澳洲', '港澳']
    },
    {
        'name': '欧美澳',
        'regions': ['北美', '欧洲', '海外其他', '亚洲', '澳洲']
    },
    {
        'name': '港澳',
        'regions': ['港澳']
    },
    {
        'name': '台湾',
        'regions': ['台湾']
    }
]


def load_config():
    env_path = Path(".env")
    if env_path.exists():
        with open(env_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#'):
                    if '=' in line:
                        key, value = line.split('=', 1)
                        os.environ[key.strip()] = value.strip()

    username = os.environ.get('SMARTBI_USERNAME')
    password = os.environ.get('SMARTBI_PASSWORD')
    base_url = os.environ.get('SMARTBI_BASE_URL', 'https://bi.61info.cn/smartbi/vision/index.jsp')

    if not username:
        username = input("请输入 Smartbi 用户名: ")
    if not password:
        password = getpass("请输入 Smartbi 密码: ")

    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

    return {
        'username': username,
        'password': password,
        'base_url': base_url,
        'download_dir': DOWNLOAD_DIR
    }


async def login(page, config):
    print("\n" + "="*60)
    print("统一登录 Smartbi (仅登录1次)")
    print("="*60)
    
    await page.goto(config['base_url'], wait_until="networkidle")

    try:
        await page.wait_for_selector('input[bofid="username"]', timeout=15000)
    except PlaywrightTimeoutError:
        raise Exception("登录页面加载超时")

    await page.fill('input[bofid="username"]', config['username'])
    await page.fill('input[bofid="password"]', config['password'])
    await page.click('input[bofid="login"]')

    await page.wait_for_load_state("networkidle", timeout=30000)
    print("[SUCCESS] 登录成功")


async def select_date_filter(page, start_date, end_date, report_type=None):
    """选择开始日期和结束日期筛选条件
    投放报表只有一个时间筛选项（结束日期），其他报表有开始日期和结束日期两个选项"""
    print(f"[INFO] 设置日期筛选: {start_date} 至 {end_date} (报表类型: {report_type})")
    
    try:
        # 判断是否为投放报表
        is_ad_report = report_type and '投放' in report_type
        
        if is_ad_report:
            # 投放报表：只有一个时间筛选项，使用结束日期
            print("[DEBUG] 投放报表，仅设置结束日期")
            
            # 方法1：查找所有input，找到日期格式的输入框
            date_input_found = False
            inputs = page.locator('input')
            count = await inputs.count()
            
            for i in range(count):
                try:
                    val = await inputs.nth(i).input_value()
                    if val and len(val) >= 10 and '-' in val:
                        print(f"[DEBUG] 找到日期输入框 {i}，当前值: {val}")
                        await inputs.nth(i).evaluate('el => el.removeAttribute("readonly")')
                        await inputs.nth(i).fill(end_date)
                        await inputs.nth(i).evaluate('el => el.dispatchEvent(new Event("change"))')
                        await inputs.nth(i).evaluate('el => el.dispatchEvent(new Event("blur"))')
                        print(f"[DEBUG] 已设置时间（结束日期）: {end_date}")
                        date_input_found = True
                        break
                except Exception as e:
                    continue
            
            if not date_input_found:
                # 尝试查找combobox-edit类
                print("[DEBUG] 尝试查找combobox-edit类输入框")
                combo_inputs = page.locator('input.combobox-edit')
                combo_count = await combo_inputs.count()
                for i in range(combo_count):
                    try:
                        await combo_inputs.nth(i).evaluate('el => el.removeAttribute("readonly")')
                        await combo_inputs.nth(i).fill(end_date)
                        await combo_inputs.nth(i).evaluate('el => el.dispatchEvent(new Event("change"))')
                        print(f"[DEBUG] 使用combobox-edit设置时间成功: {end_date}")
                        date_input_found = True
                        break
                    except Exception as e:
                        continue
            
            if not date_input_found:
                print(f"[WARNING] 未找到日期输入框")
            
        else:
            # 非投放报表：有开始日期和结束日期两个选项
            print("[DEBUG] 商务报表，设置开始日期和结束日期")
            
            # 查找所有日期输入框
            date_inputs = []
            inputs = page.locator('input')
            count = await inputs.count()
            
            for i in range(count):
                try:
                    val = await inputs.nth(i).input_value()
                    if val and len(val) >= 10 and '-' in val:
                        date_inputs.append(i)
                        print(f"[DEBUG] 找到日期输入框 {i}，当前值: {val}")
                except Exception as e:
                    continue
            
            if len(date_inputs) >= 2:
                # 根据页面标签判断输入框顺序
                body_text = await page.evaluate('document.body.textContent || ""')
                start_pos = body_text.find('开始日期')
                end_pos = body_text.find('结束日期')
                
                print(f"[DEBUG] 开始日期标签位置: {start_pos}, 结束日期标签位置: {end_pos}")
                
                if start_pos >= 0 and end_pos >= 0:
                    if start_pos < end_pos:
                        start_idx, end_idx = date_inputs[0], date_inputs[1]
                    else:
                        start_idx, end_idx = date_inputs[1], date_inputs[0]
                else:
                    # 默认第一个是结束日期，第二个是开始日期
                    start_idx, end_idx = date_inputs[1], date_inputs[0]
                
                print(f"[DEBUG] 开始日期输入框索引: {start_idx}, 结束日期输入框索引: {end_idx}")
                
                # 设置开始日期
                await inputs.nth(start_idx).evaluate('el => el.removeAttribute("readonly")')
                await inputs.nth(start_idx).fill(start_date)
                await inputs.nth(start_idx).evaluate('el => el.dispatchEvent(new Event("change"))')
                await inputs.nth(start_idx).evaluate('el => el.dispatchEvent(new Event("blur"))')
                
                # 设置结束日期
                await inputs.nth(end_idx).evaluate('el => el.removeAttribute("readonly")')
                await inputs.nth(end_idx).fill(end_date)
                await inputs.nth(end_idx).evaluate('el => el.dispatchEvent(new Event("change"))')
                await inputs.nth(end_idx).evaluate('el => el.dispatchEvent(new Event("blur"))')
                
                print(f"[DEBUG] 已设置日期范围: {start_date} 至 {end_date}")
            else:
                print(f"[WARNING] 未找到足够的日期输入框，找到: {len(date_inputs)} 个")
            
        # 尝试点击查询按钮应用筛选条件
        print("[DEBUG] 尝试点击查询按钮应用筛选")
        query_applied = False
        
        # 方法1：查找包含特定文本的按钮
        query_buttons = ['查询', '搜索', '确定', '刷新', '应用', '检索']
        for btn_text in query_buttons:
            try:
                btn = page.locator(f'button:has-text("{btn_text}")')
                if await btn.count() > 0:
                    await btn.click()
                    print(f"[DEBUG] 点击了 '{btn_text}' 按钮")
                    await page.wait_for_load_state("networkidle", timeout=30000)
                    query_applied = True
                    break
            except Exception as e:
                continue
        
        if not query_applied:
            try:
                icons = page.locator('button[class*="icon"]')
                if await icons.count() > 0:
                    for i in range(await icons.count()):
                        try:
                            await icons.nth(i).click()
                            print(f"[DEBUG] 点击了图标按钮 {i}")
                            await page.wait_for_load_state("networkidle", timeout=30000)
                            query_applied = True
                            break
                        except:
                            continue
            except Exception as e:
                print(f"[DEBUG] 查找图标按钮失败: {e}")
        
        if not query_applied:
            try:
                refresh_icons = page.locator('[title="刷新"], [title="查询"], [title="确定"]')
                if await refresh_icons.count() > 0:
                    await refresh_icons.first.click()
                    print("[DEBUG] 点击了带title属性的刷新/查询按钮")
                    await page.wait_for_load_state("networkidle", timeout=30000)
                    query_applied = True
            except Exception as e:
                print(f"[DEBUG] 查找带title属性的按钮失败: {e}")
        
        if not query_applied:
            try:
                # 方法4：查找带有特定class的按钮
                action_buttons = page.locator('.btn, .button, .search-btn, .query-btn, .refresh-btn')
                if await action_buttons.count() > 0:
                    for i in range(await action_buttons.count()):
                        btn = action_buttons.nth(i)
                        try:
                            await btn.click()
                            print(f"[DEBUG] 点击了class选择的按钮 {i}")
                            await page.wait_for_load_state("networkidle", timeout=30000)
                            query_applied = True
                            break
                        except:
                            continue
            except Exception as e:
                print(f"[DEBUG] 方法4 - 查找class按钮失败: {e}")
        
        if query_applied:
            print("[DEBUG] 成功应用日期筛选条件")
        else:
            print("[DEBUG] 未找到查询按钮，可能页面会自动刷新")
        
        await asyncio.sleep(3)
        
    except Exception as e:
        print(f"[WARNING] 设置日期筛选失败: {e}")


async def select_region_filter(page, regions):
    """选择区域等级筛选条件 - 直接在输入框中输入区域选项"""
    print(f"[INFO] 设置区域筛选: {regions}")
    
    try:
        await page.wait_for_selector('.combobox-panel', timeout=15000)
        panels = page.locator('.combobox-panel')
        count = await panels.count()
        print(f"[DEBUG] 找到 {count} 个 combobox-panel")
        
        target_panel = None
        target_input = None
        
        for i in range(count):
            panel = panels.nth(i)
            title = await panel.get_attribute('title')
            print(f"[DEBUG] Panel {i} title: {title}")
            
            if title and ('欧洲' in title or '北美' in title or '亚洲' in title or '港澳' in title or '台湾' in title):
                target_panel = panel
                target_input = panel.locator('input.combobox-edit')
                break
        
        if not target_panel or not target_input:
            print("[DEBUG] 方法一未找到，尝试方法二：查找区域等级标签附近的输入框")
            try:
                region_label = page.locator('td:has-text("区域等级")')
                if await region_label.count() > 0:
                    print("[DEBUG] 找到区域等级标签")
                    parent_tr = region_label.locator('..')
                    input_el = parent_tr.locator('input.combobox-edit')
                    if await input_el.count() > 0:
                        target_input = input_el
            except Exception as e:
                print(f"[DEBUG] 方法二失败: {e}")
        
        if not target_input:
            print("[DEBUG] 尝试方法三：查找所有combobox-edit输入框")
            inputs = page.locator('input.combobox-edit')
            input_count = await inputs.count()
            print(f"[DEBUG] 找到 {input_count} 个 combobox-edit 输入框")
            
            for i in range(input_count):
                input_el = inputs.nth(i)
                value = await input_el.input_value()
                print(f"[DEBUG] 输入框 {i} 值: {value}")
                if value and ('全部' in value or '亚洲' in value or '欧洲' in value):
                    target_input = input_el
                    break
        
        if target_input:
            await target_input.click()
            await target_input.clear()
            
            # 输入区域（用逗号分隔）
            region_str = ','.join(regions)
            await target_input.type(region_str)
            print(f"[DEBUG] 已输入区域: {region_str}")
            
            await asyncio.sleep(1)
            
            # 尝试触发下拉选择
            try:
                await page.keyboard.press('Enter')
            except:
                pass
        else:
            print("[WARNING] 未找到区域筛选输入框")
            
    except Exception as e:
        print(f"[ERROR] 设置区域筛选失败: {e}")


async def export_report(page, config, report_name, region_name):
    """导出报表 - 使用原始的导出流程"""
    print(f"[INFO] 开始导出: {report_name} - {region_name}")
    
    download_file_path = None
    
    async def handle_download(download):
        nonlocal download_file_path
        file_name = download.suggested_filename or f"report_{int(time.time())}.xlsx"
        date_suffix = time.strftime("%y%m%d")
        name_parts = file_name.rsplit('.', 1)
        
        if len(name_parts) > 1:
            # 添加区域名称到文件名
            base_name = f"{name_parts[0]}_{date_suffix}_{region_name}"
            extension = name_parts[1]
        else:
            base_name = f"{file_name}_{date_suffix}_{region_name}"
            extension = "xlsx"
        
        download_file_path = DOWNLOAD_DIR / f"{base_name}.{extension}"
        counter = 1
        while download_file_path.exists():
            download_file_path = DOWNLOAD_DIR / f"{base_name}_v{counter}.{extension}"
            counter += 1
        
        await download.save_as(str(download_file_path))
        print(f"[DEBUG] 文件下载完成: {download_file_path.name}")
    
    # 等待页面加载完成
    await page.wait_for_load_state("networkidle", timeout=60000)
    
    # 注册下载事件监听器（保存handle_download函数引用用于移除）
    page.on("download", handle_download)
    
    try:
        # 点击导出按钮（使用原始的选择器）
        print("[DEBUG] 点击导出按钮")
        export_selectors = ['input:has-text("导出")', 'button:has-text("导出")', '[class*="export"]']
        export_found = False
        for selector in export_selectors:
            if await page.locator(selector).count() > 0:
                await page.click(selector)
                export_found = True
                print(f"[DEBUG] 找到导出按钮: {selector}")
                break
        
        if not export_found:
            raise Exception("未找到导出按钮")
        
        # 选择 Excel 格式
        print("[DEBUG] 选择 Excel 格式")
        try:
            await page.wait_for_selector('span:has-text("Excel")', timeout=5000)
            await page.click('span:has-text("Excel")')
        except:
            # 备用方法：查找包含Excel的元素
            excel_options = page.locator('*:has-text("Excel")')
            if await excel_options.count() > 0:
                await excel_options.first.click()
                print("[DEBUG] 备用方法选择Excel格式")
            else:
                raise Exception("未找到 Excel 选项")
        
        # 点击在线导出按钮
        print("[DEBUG] 等待在线导出按钮")
        online_export_selector = 'input[value="在线导出"]'
        
        try:
            await page.wait_for_selector(online_export_selector, timeout=10000)
            await page.click(online_export_selector)
            print("[DEBUG] 已点击在线导出")
        except:
            try:
                await page.wait_for_selector('input[atp="baseDialog_btnOK"]', timeout=5000)
                await page.click('input[atp="baseDialog_btnOK"]')
                print("[DEBUG] 已点击在线导出(备用选择器)")
            except:
                raise Exception("未找到在线导出按钮")
        
        # 等待下载完成
        print("[INFO] 等待下载完成...")
        download_timeout = 180
        start_time = time.time()
        
        while time.time() - start_time < download_timeout:
            if download_file_path and download_file_path.exists():
                break
            await asyncio.sleep(1)
            if int(time.time() - start_time) % 30 == 0:
                print(f"[INFO] 等待中... ({int(time.time() - start_time)}秒)")
        
        if not download_file_path or not download_file_path.exists():
            raise Exception(f"下载超时！{report_name} - {region_name} 报表未成功下载")
        
        print(f"[SUCCESS] {report_name} - {region_name} 导出完成")
        return str(download_file_path)
    
    finally:
        # 移除下载事件监听器
        try:
            page.remove_listener("download", handle_download)
        except Exception as e:
            print(f"[DEBUG] 移除下载监听器时发生错误（可忽略）: {e}")


# ============ 主投ROI2数据提取相关函数 ============

def parse_report_name(file_name):
    """解析报表名称和区域（去除区域名称后的版本号如v2、v3等）"""
    pattern = r'(.+)_(\d{6})_(.+)\.xlsx'
    match = re.match(pattern, file_name)
    if match:
        region = match.group(3)
        # 去除区域名称后的版本号（如 v2, v3, v1 等）
        region = re.sub(r'v\d+$', '', region)
        # 去除末尾可能的下划线
        region = region.rstrip('_')
        return {
            'original_name': match.group(1),
            'date': match.group(2),
            'region': region
        }
    return None


def extract_report_date(ws):
    """提取报表中的日期参数
    - 投放报表：提取"时间"字段的值
    - 商务报表：提取"结束日期"字段的值（用于比对）"""
    result = None
    
    for row in range(1, min(15, ws.max_row + 1)):
        for col in range(1, min(15, ws.max_column + 1)):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                val = str(cell.value).strip()
                # 优先查找"结束日期"或"时间"
                if val == '结束日期':
                    next_cell = ws.cell(row=row, column=col+1)
                    if next_cell.value:
                        date_val = str(next_cell.value).strip()[:10]
                        print(f"[DEBUG] 找到结束日期: {date_val}")
                        return date_val
                elif val == '时间':
                    next_cell = ws.cell(row=row, column=col+1)
                    if next_cell.value:
                        date_val = str(next_cell.value).strip()[:10]
                        print(f"[DEBUG] 找到时间: {date_val}")
                        return date_val
    
    # 如果没找到"结束日期"或"时间"，再查找包含"日期"的字段
    for row in range(1, min(15, ws.max_row + 1)):
        for col in range(1, min(15, ws.max_column + 1)):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                val = str(cell.value).strip()
                if '日期' in val:
                    next_cell = ws.cell(row=row, column=col+1)
                    if next_cell.value:
                        date_val = str(next_cell.value).strip()[:10]
                        print(f"[DEBUG] 找到日期字段: {val} -> {date_val}")
                        return date_val
    
    return None


def extract_roi2_data(ws, report_type):
    """提取主投=思维，二级渠道/供应商=汇总/总计的滚动ROI2达成数据"""
    
    # 根据报表类型确定目标列
    # 投放报表: I列(9) - 主投-滚动ROI2达成
    # 非港澳商务(欧美澳): H列(8) - 主投-滚动ROI2达成
    # 港澳商务、台湾商务、Local商务: I列(9) - 主投-滚动ROI2达成
    if report_type == '非港澳商务':
        target_col = 8  # H列
    else:
        target_col = 9  # I列
    
    print(f"[DEBUG] {report_type} - 目标列: {target_col}, 最大列数: {ws.max_column}")
    
    # 查找主投=思维的起始行
    thought_start_row = None
    for row in range(1, min(200, ws.max_row + 1)):
        cell_b = ws.cell(row=row, column=2)
        if cell_b.value:
            cell_val = str(cell_b.value).strip()
            if cell_val == '思维':
                thought_start_row = row
                print(f"[DEBUG] {report_type} - 找到主投=思维的行: {row}")
                break
            # 打印一些候选行用于调试
            if '思' in cell_val or '投' in cell_val:
                print(f"[DEBUG] {report_type} - 行{row} B列值: '{cell_val}'")
    
    if thought_start_row is None:
        print(f"[DEBUG] {report_type} - 未找到主投=思维的行")
        # 打印前20行B列的值用于调试
        print(f"[DEBUG] {report_type} - 前20行B列值:")
        for row in range(1, min(21, ws.max_row + 1)):
            cell_b = ws.cell(row=row, column=2)
            if cell_b.value:
                print(f"[DEBUG]   行{row}: '{str(cell_b.value).strip()}'")
        return None
    
    # 根据报表类型提取数据
    if report_type == '投放':
        # 投放报表：主投=思维 AND 辅投=汇总 AND 二级渠道=汇总
        # B列=主投类型，C列=辅投，D列=二级渠道
        # 处理合并单元格：跟踪当前主投类型
        print(f"[DEBUG] 投放报表 - 搜索范围: 行{thought_start_row} 到 行{min(thought_start_row + 50, ws.max_row + 1)}")
        
        current_main_type = None
        for row in range(thought_start_row, min(thought_start_row + 50, ws.max_row + 1)):
            cell_b = ws.cell(row=row, column=2)
            cell_c = ws.cell(row=row, column=3)
            cell_d = ws.cell(row=row, column=4)
            
            b_val = str(cell_b.value).strip() if cell_b.value else ''
            c_val = str(cell_c.value).strip() if cell_c.value else ''
            d_val = str(cell_d.value).strip() if cell_d.value else ''
            
            # 更新当前主投类型（处理合并单元格）
            if b_val != '':
                current_main_type = b_val
            
            # 打印思维区块的所有行
            if current_main_type == '思维':
                print(f"[DEBUG] 投放报表 - 行{row}: current_main_type='{current_main_type}', C='{c_val}', D='{d_val}'")
            
            # 检查是否已离开思维区块
            if current_main_type not in ['思维', None]:
                break
            
            # 在思维区块内查找辅投=汇总 AND 二级渠道=汇总的行
            if current_main_type == '思维' and c_val == '汇总' and d_val == '汇总':
                if target_col <= ws.max_column:
                    roi2_cell = ws.cell(row=row, column=target_col)
                    if roi2_cell.value is not None and isinstance(roi2_cell.value, (int, float)):
                        print(f"[DEBUG] 投放报表 - 找到数据: 行{row}, 值={roi2_cell.value}")
                        return roi2_cell.value
    
    elif report_type == '非港澳商务':
        # 非港澳商务：主投=思维 AND 二级渠道=总计
        # B列=主投类型，C列=二级渠道
        # 处理合并单元格：跟踪当前主投类型，遇到新类型时更新
        print(f"[DEBUG] 非港澳商务 - 搜索范围: 行{thought_start_row} 到 行{min(thought_start_row + 50, ws.max_row + 1)}")
        
        current_main_type = None
        for row in range(thought_start_row, min(thought_start_row + 50, ws.max_row + 1)):
            cell_b = ws.cell(row=row, column=2)
            cell_c = ws.cell(row=row, column=3)
            
            b_val = str(cell_b.value).strip() if cell_b.value else ''
            c_val = str(cell_c.value).strip() if cell_c.value else ''
            
            # 更新当前主投类型（处理合并单元格）
            if b_val != '':
                current_main_type = b_val
            
            # 打印思维区块的所有行
            if current_main_type == '思维':
                print(f"[DEBUG] 非港澳商务 - 行{row}: current_main_type='{current_main_type}', C='{c_val}'")
            
            # 在思维区块内查找二级渠道=总计的行
            if current_main_type == '思维' and c_val == '总计':
                if target_col <= ws.max_column:
                    roi2_cell = ws.cell(row=row, column=target_col)
                    if roi2_cell.value is not None and isinstance(roi2_cell.value, (int, float)):
                        print(f"[DEBUG] 非港澳商务 - 找到数据: 行{row}, 值={roi2_cell.value}")
                        return roi2_cell.value
    
    elif report_type == '港澳商务':
        # 港澳商务：主投=思维 AND 二级渠道=总计 AND 供应商=总计
        # B列=主投类型，C列=二级渠道，D列=供应商
        # 处理合并单元格：跟踪当前主投类型
        current_main_type = None
        for row in range(thought_start_row, min(thought_start_row + 100, ws.max_row + 1)):
            cell_b = ws.cell(row=row, column=2)
            cell_c = ws.cell(row=row, column=3)
            cell_d = ws.cell(row=row, column=4)
            
            b_val = str(cell_b.value).strip() if cell_b.value else ''
            c_val = str(cell_c.value).strip() if cell_c.value else ''
            d_val = str(cell_d.value).strip() if cell_d.value else ''
            
            # 更新当前主投类型（处理合并单元格）
            if b_val != '':
                current_main_type = b_val
            
            # 检查是否已离开思维区块
            if current_main_type not in ['思维', None]:
                break
            
            if current_main_type == '思维' and c_val == '总计' and d_val == '总计':
                if target_col <= ws.max_column:
                    roi2_cell = ws.cell(row=row, column=target_col)
                    if roi2_cell.value is not None and isinstance(roi2_cell.value, (int, float)):
                        print(f"[DEBUG] 港澳商务 - 找到数据: 行{row}, 值={roi2_cell.value}")
                        return roi2_cell.value
    
    elif report_type == '台湾商务':
        # 台湾商务：主投=思维 AND 二级渠道=台湾商务-总计 AND 供应商=总计
        # B列=主投类型，C列=二级渠道，D列=供应商
        # 处理合并单元格：跟踪当前主投类型
        current_main_type = None
        for row in range(thought_start_row, min(thought_start_row + 100, ws.max_row + 1)):
            cell_b = ws.cell(row=row, column=2)
            cell_c = ws.cell(row=row, column=3)
            cell_d = ws.cell(row=row, column=4)
            
            b_val = str(cell_b.value).strip() if cell_b.value else ''
            c_val = str(cell_c.value).strip() if cell_c.value else ''
            d_val = str(cell_d.value).strip() if cell_d.value else ''
            
            # 更新当前主投类型（处理合并单元格）
            if b_val != '':
                current_main_type = b_val
            
            # 检查是否已离开思维区块
            if current_main_type not in ['思维', None]:
                break
            
            if current_main_type == '思维' and c_val == '台湾商务-总计' and d_val == '总计':
                if target_col <= ws.max_column:
                    roi2_cell = ws.cell(row=row, column=target_col)
                    if roi2_cell.value is not None and isinstance(roi2_cell.value, (int, float)):
                        print(f"[DEBUG] 台湾商务 - 找到数据: 行{row}, 值={roi2_cell.value}")
                        return roi2_cell.value
    
    elif report_type == 'Local商务':
        # Local商务：主投=思维 AND 二级渠道=总计 AND 供应商=总计
        # B列=主投类型，C列=二级渠道，D列=供应商
        # 处理合并单元格：跟踪当前主投类型
        current_main_type = None
        for row in range(thought_start_row, min(thought_start_row + 100, ws.max_row + 1)):
            cell_b = ws.cell(row=row, column=2)
            cell_c = ws.cell(row=row, column=3)
            cell_d = ws.cell(row=row, column=4)
            
            b_val = str(cell_b.value).strip() if cell_b.value else ''
            c_val = str(cell_c.value).strip() if cell_c.value else ''
            d_val = str(cell_d.value).strip() if cell_d.value else ''
            
            # 更新当前主投类型（处理合并单元格）
            if b_val != '':
                current_main_type = b_val
            
            # 检查是否已离开思维区块
            if current_main_type not in ['思维', None]:
                break
            
            if current_main_type == '思维' and c_val == '总计' and d_val == '总计':
                if target_col <= ws.max_column:
                    roi2_cell = ws.cell(row=row, column=target_col)
                    if roi2_cell.value is not None and isinstance(roi2_cell.value, (int, float)):
                        print(f"[DEBUG] Local商务 - 找到数据: 行{row}, 值={roi2_cell.value}")
                        return roi2_cell.value
    
    print(f"[DEBUG] {report_type} - 未找到匹配的数据")
    return None


def extract_all_roi2_data(downloaded_files=None):
    """提取下载报表中的主投ROI2数据
    只处理本次下载的文件列表，不会回退到扫描整个目录"""
    results = []
    all_dates = set()
    
    # 如果传入了下载文件列表，只处理这些文件
    if downloaded_files is not None:
        # 过滤掉None值、空字符串和无效路径
        valid_files = []
        for f in downloaded_files:
            if f:
                str_path = str(f).strip()
                if str_path:
                    valid_files.append(Path(str_path))
        
        if not valid_files:
            print("[WARNING] 本次下载文件列表为空，跳过ROI2数据提取")
            return results, all_dates
        
        files_to_process = valid_files
        print(f"[DEBUG] 只处理本次下载的 {len(files_to_process)} 个文件")
        print(f"[DEBUG] 下载文件列表: {[str(f) for f in files_to_process]}")
    else:
        # 未传入下载文件列表，直接返回空结果
        print("[WARNING] 未传入下载文件列表，跳过ROI2数据提取")
        return results, all_dates
    
    for file_path in files_to_process:
        print(f"[DEBUG] 处理文件: {file_path.name}")
        parsed = parse_report_name(file_path.name)
        if not parsed:
            print(f"[DEBUG] 文件解析失败: {file_path.name}")
            continue
        
        print(f"[DEBUG] 解析结果: original_name={parsed['original_name']}, date={parsed['date']}, region={parsed['region']}")
        
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            print(f"[DEBUG] 工作表名称: {ws.title}, 行数: {ws.max_row}, 列数: {ws.max_column}")
            
            # 判断报表类型
            original_name = parsed['original_name']
            if '投放' in original_name:
                report_type = '投放'
            elif '非港澳商务' in original_name:
                report_type = '非港澳商务'
            elif '港澳商务' in original_name:
                report_type = '港澳商务'
            elif '台湾商务' in original_name:
                report_type = '台湾商务'
            elif 'Local商务' in original_name:
                report_type = 'Local商务'
            else:
                report_type = '非港澳商务'
            
            # 提取ROI2数据
            roi2 = extract_roi2_data(ws, report_type)
            
            # 提取报表日期
            report_date = extract_report_date(ws)
            if report_date:
                all_dates.add(report_date)
            
            print(f"[DEBUG] {report_type} - {parsed['region']} - ROI2值: {roi2}")
            
            results.append({
                'report_type': report_type,
                'region': parsed['region'],
                'report_date': report_date,
                'roi2': roi2
            })
            
            wb.close()
        except Exception as e:
            print(f"[ERROR] 读取文件 {file_path.name} 失败: {e}")
    
    return results, all_dates


def create_roi2_report(roi2_data, report_dates, target_column='更新前'):
    """生成主投ROI2数据比对Excel文件"""
    today = time.strftime("%y%m%d")
    output_file = OUTPUT_DIR / f"主投ROI2数据比对_{today}.xlsx"
    
    # 检查日期一致性
    if len(report_dates) > 1:
        print(f"[WARNING] 发现多个不同日期: {report_dates}")
        confirm = input("日期不一致，是否继续执行？(y/n): ")
        if confirm.lower() != 'y':
            print("已取消执行")
            return
    
    data_date = list(report_dates)[0] if report_dates else time.strftime("%Y-%m-%d")
    update_time = time.strftime("%Y-%m-%d %H:%M:%S")
    
    # 按报表类型和区域组织数据
    roi2_map = {}
    for item in roi2_data:
        key = (item['report_type'], item['region'])
        roi2_map[key] = item['roi2']
    
    # 确定目标列索引
    # A:报表类型, B:区域, C:更新前, D:更新后, E:gap
    target_col_idx = 3 if target_column == '更新前' else 4
    
    # 确定是否需要创建新文件或添加新表
    file_exists = output_file.exists()
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
    
    if file_exists:
        wb = load_workbook(output_file)
        ws = wb.active
        
        # 查找已有数据的行范围（最新的表格）
        last_table_start = 2
        for row in range(max(1, ws.max_row - 50), ws.max_row + 1):
            cell_a = ws.cell(row=row, column=1)
            cell_b = ws.cell(row=row, column=2)
            cell_c = ws.cell(row=row, column=3)
            if cell_a.value == '报表类型' and cell_b.value == '区域' and cell_c.value == '更新前':
                last_table_start = row
                break
        
        # 在最新表格中查找匹配的行
        updated_keys = set()
        for row in range(last_table_start + 1, ws.max_row + 1):
            cell_a = ws.cell(row=row, column=1)
            cell_b = ws.cell(row=row, column=2)
            cell_target = ws.cell(row=row, column=target_col_idx)
            
            report_type = str(cell_a.value).strip() if cell_a.value else ''
            region = str(cell_b.value).strip() if cell_b.value else ''
            
            if report_type and region:
                key = (report_type, region)
                if key in roi2_map and cell_target.value is None:
                    roi2_val = roi2_map[key]
                    if roi2_val is not None:
                        cell_target.value = roi2_val
                        cell_target.number_format = '0.000'
                        cell_target.border = thin_border
                    updated_keys.add(key)
        
        # 检查是否有未更新的数据
        remaining_keys = roi2_map.keys() - updated_keys
        
        if remaining_keys:
            current_row = ws.max_row + 1
            
            for report_type in ['投放', '非港澳商务', '港澳商务', '台湾商务', 'Local商务']:
                for region in list(set(k[1] for k in remaining_keys if k[0] == report_type)):
                    key = (report_type, region)
                    if key in roi2_map and key in remaining_keys:
                        ws.cell(row=current_row, column=1, value=report_type).border = thin_border
                        ws.cell(row=current_row, column=2, value=region).border = thin_border
                        
                        roi2_val = roi2_map[key]
                        if roi2_val is not None:
                            cell = ws.cell(row=current_row, column=target_col_idx, value=roi2_val)
                            cell.number_format = '0.000'
                            cell.border = thin_border
                        
                        gap_cell = ws.cell(row=current_row, column=5, value=f"=D{current_row}-C{current_row}")
                        gap_cell.number_format = '0.000'
                        gap_cell.border = thin_border
                        
                        current_row += 1
        
        print(f"[INFO] 找到已存在的文件，将更新'{target_column}'列")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "主投ROI2数据比对"
        
        # 添加时间信息
        ws.merge_cells('A1:E1')
        time_cell = ws.cell(row=1, column=1, value=f"数据更新时间: {update_time} | 数据时间: {data_date}")
        time_cell.font = Font(bold=True)
        time_cell.alignment = Alignment(horizontal='center')
        
        # 创建表头
        headers = ['报表类型', '区域', '更新前', '更新后', 'gap']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # 设置列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        
        current_row = 3
        
        # 按报表类型和区域组织要显示的数据
        report_types = list(set(item['report_type'] for item in roi2_data))
        report_types.sort(key=['投放', '非港澳商务', '港澳商务', '台湾商务', 'Local商务'].index)
        
        regions = list(set(item['region'] for item in roi2_data))
        
        # 填充数据
        for report_type in report_types:
            for region in regions:
                key = (report_type, region)
                if key in roi2_map:
                    ws.cell(row=current_row, column=1, value=report_type).border = thin_border
                    ws.cell(row=current_row, column=2, value=region).border = thin_border
                    
                    roi2_val = roi2_map[key]
                    if roi2_val is not None:
                        cell = ws.cell(row=current_row, column=target_col_idx, value=roi2_val)
                        cell.number_format = '0.000'
                        cell.border = thin_border
                    
                    gap_cell = ws.cell(row=current_row, column=5, value=f"=D{current_row}-C{current_row}")
                    gap_cell.number_format = '0.000'
                    gap_cell.border = thin_border
                    
                    current_row += 1
    
    wb.save(output_file)
    print(f"[SUCCESS] 主投ROI2数据比对文件已生成: {output_file}")


async def main():
    print("=" * 70)
    print("商务报表批量导出脚本")
    print("=" * 70)

    default_reports = [r for r in REPORT_CONFIGS if r.get('default', False)]
    
    print("\n可用的报表:")
    for i, report in enumerate(REPORT_CONFIGS):
        default_marker = "*" if report.get('default', False) else ""
        print(f"  {i+1}. {report['name']}{default_marker}")
    print("  (*表示默认导出)")

    selected_report_indices = input("\n请输入要导出的报表序号（多个用逗号分隔，默认导出标*的报表）: ")
    
    if selected_report_indices.strip():
        parts = selected_report_indices.replace('，', ',').split(',')
        selected_report_indices = [int(i.strip()) - 1 for i in parts if i.strip().isdigit()]
        selected_reports = [(i, REPORT_CONFIGS[i]) for i in selected_report_indices if 0 <= i < len(REPORT_CONFIGS)]
    else:
        selected_reports = [(i, r) for i, r in enumerate(REPORT_CONFIGS) if r.get('default', False)]

    print("\n可用的区域等级筛选选项:")
    for i, config in enumerate(REGION_CONFIGS):
        print(f"  {i+1}. {config['name']}")
        print(f"     包含区域: {', '.join(config['regions'])}")

    print("\n" + "-" * 50)
    print("请为每个报表选择区域筛选（直接回车使用默认配置）:")
    print("-" * 50)
    
    report_region_map = {}
    for idx, report in selected_reports:
        default_region_indices = report.get('default_regions', [0])
        default_region_names = [REGION_CONFIGS[i]['name'] for i in default_region_indices]
        default_str = ','.join(str(i+1) for i in default_region_indices)
        
        print(f"\n{report['name']}:")
        print(f"  默认配置: {', '.join(default_region_names)}")
        user_input = input(f"  请输入区域序号（多个用逗号分隔，默认{default_str}）: ").strip()
        
        if user_input:
            parts = user_input.replace('，', ',').split(',')
            region_indices = [int(i.strip()) - 1 for i in parts if i.strip().isdigit()]
            selected_regions = [REGION_CONFIGS[i] for i in region_indices if 0 <= i < len(REGION_CONFIGS)]
        else:
            selected_regions = [REGION_CONFIGS[i] for i in default_region_indices if 0 <= i < len(REGION_CONFIGS)]
        
        report_region_map[idx] = selected_regions

    print(f"\n将依次处理以下报表及区域:")
    for idx, report in selected_reports:
        regions = report_region_map[idx]
        region_names = [r['name'] for r in regions]
        print(f"  - {report['name']} (idx={idx}): {', '.join(region_names)}")

    # 询问用户日期筛选
    print("\n日期筛选设置:")
    
    # 计算默认日期
    today = datetime.date.today()
    
    # 开始日期默认值：如果今天是1号，默认上个月1号；否则默认当月1号
    if today.day == 1:
        # 今天是1号，返回上个月1号
        if today.month == 1:
            default_start_date = datetime.date(today.year - 1, 12, 1)
        else:
            default_start_date = datetime.date(today.year, today.month - 1, 1)
    else:
        # 今天不是1号，返回当月1号
        default_start_date = datetime.date(today.year, today.month, 1)
    
    # 结束日期默认值：昨天
    default_end_date = today - datetime.timedelta(days=1)
    
    start_date_str = default_start_date.strftime("%Y-%m-%d")
    end_date_str = default_end_date.strftime("%Y-%m-%d")
    
    start_date = input(f"请输入开始日期（格式：YYYY-MM-DD，默认{start_date_str}）: ").strip()
    if not start_date:
        start_date = start_date_str
    
    end_date = input(f"请输入结束日期（格式：YYYY-MM-DD，默认{end_date_str}）: ").strip()
    if not end_date:
        end_date = end_date_str
    
    print(f"[INFO] 将筛选日期范围: {start_date} 至 {end_date}")

    # 询问用户数据填入哪一列
    print("\n数据填写位置选择:")
    print("  1. 更新前")
    print("  2. 更新后")
    while True:
        col_choice = input("请选择数据填入哪一列（1/2，默认1）: ").strip()
        if col_choice == '':
            target_column = '更新前'
            break
        elif col_choice == '1':
            target_column = '更新前'
            break
        elif col_choice == '2':
            target_column = '更新后'
            break
        else:
            print("请输入有效选项 (1 或 2)")
    
    print(f"\n[INFO] 本次数据将填入 '{target_column}' 列")

    config = load_config()

    playwright = await async_playwright().start()
    browser = await playwright.chromium.launch(
        headless=False,
        args=["--start-maximized", "--disable-popup-blocking"]
    )
    context = await browser.new_context(viewport={"width": 1920, "height": 1080})
    page = await context.new_page()
    
    page.on("dialog", lambda dialog: dialog.accept())

    downloaded_files = []
    try:
        await login(page, config)

        for idx, report_config in selected_reports:
            selected_regions = report_region_map[idx]
            
            print(f"\n{'='*60}")
            print(f"访问报表: {report_config['name']}")
            print(f"{'='*60}")
            print(f"[DEBUG] 该报表将导出 {len(selected_regions)} 个区域")
            
            await page.goto(report_config['url'], wait_until="networkidle")
            
            try:
                await page.wait_for_selector('table', timeout=60000)
            except:
                pass
            
            await page.wait_for_load_state("networkidle", timeout=60000)
            print(f"[INFO] {report_config['name']} 报表页面加载完成")

            # 先设置日期筛选（每个报表只设置一次）
            print(f"[INFO] 设置日期筛选: {start_date} 至 {end_date}")
            await select_date_filter(page, start_date, end_date, report_config['name'])
            
            for i, region_config in enumerate(selected_regions):
                try:
                    print(f"\n{'='*50}")
                    print(f"处理 [{i+1}/{len(selected_regions)}]: {report_config['name']} - {region_config['name']}")
                    print(f"{'='*50}")
                    print(f"[DEBUG] 区域配置: {region_config}")

                    await select_region_filter(page, region_config['regions'])
                    
                    await asyncio.sleep(2)

                    download_path = await export_report(page, config, report_config['name'], region_config['name'])
                    
                    # 只有成功下载的文件才添加到列表中
                    if download_path and str(download_path).strip():
                        downloaded_files.append(download_path)
                        print(f"[DEBUG] 添加下载文件: {download_path}")
                    else:
                        print(f"[WARNING] 导出返回空路径，跳过添加")

                    await asyncio.sleep(2)

                except Exception as e:
                    print(f"[ERROR] {report_config['name']} - {region_config['name']} 导出失败: {e}")

    finally:
        await context.close()
        await browser.close()

    # 提取ROI2数据并生成比对报表
    print("\n" + "="*60)
    print("提取主投ROI2数据")
    print("="*60)
    
    roi2_data, report_dates = extract_all_roi2_data(downloaded_files)
    
    if roi2_data:
        print("\n提取的数据:")
        for item in roi2_data:
            print(f"  {item['report_type']} - {item['region']}: ROI2={item['roi2']}")
        
        create_roi2_report(roi2_data, report_dates, target_column)
    else:
        print("[WARNING] 未提取到任何ROI2数据")


if __name__ == "__main__":
    asyncio.run(main())
