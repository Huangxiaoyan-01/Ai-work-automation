import asyncio
import os
import shutil
import time
import argparse
import traceback
from pathlib import Path
from getpass import getpass

from playwright.async_api import Playwright, async_playwright, Dialog, TimeoutError as PlaywrightTimeoutError
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import CellRichText, TextBlock

BASE_DIR = Path(__file__).resolve().parent

DOWNLOAD_DIR = Path(r"C:\Users\huangxiaoyan\Documents\trae_projects\test_week\downloads")
OUTPUT_DIR = Path(r"C:\Users\huangxiaoyan\Documents\trae_projects\test_week\output")
SCREENSHOT_DIR = Path(r"C:\Users\huangxiaoyan\Documents\trae_projects\test_week\screenshots")

GREEN_COLOR = "2EA121"
RED_COLOR = "F54A45"
FEISHU_DEFAULT_FONT = "Microsoft YaHei"

BUSINESS_CONFIGS = {
    '欧美澳商务': {
        'report_url': 'https://bi.61info.cn/smartbi/vision/openresource.jsp?resid=I2c928087019dd055d0552dd4019ddd0eefea7ba7',
        'title': '1）欧美澳商务',
        'extract_channels': True,
        'channel_type': 'community',  # 欧美澳商务提取的是"社群"渠道
        'channel_name': '社群',
    },
    '港澳商务': {
        'report_url': 'https://bi.61info.cn/smartbi/vision/openresource.jsp?resid=I2c928087019dd055d0552dd4019ddd3d366c3e85',
        'title': '2）港澳商务',
        'extract_channels': True,
        'channel_type': 'dual',  # 港澳商务提取的是"KOLHK"和"线下HK"两个渠道
        'kol_channel': 'KOLHK',
        'offline_channel': '线下HK',
    },
    'Local商务': {
        'report_url': 'https://bi.61info.cn/smartbi/vision/openresource.jsp?resid=I2c928087019dd055d0552dd4019ddd3d94a83eee',
        'title': '3）Local商务',
        'extract_channels': False,
    }
}

def load_config():
    env_path = BASE_DIR / ".env"
    if env_path.exists():
        with open(env_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#'):
                    if '=' in line:
                        key, value = line.split('=', 1)
                        os.environ[key.strip()] = value.strip()

    username = os.environ.get('SMARTBI_USERNAME')
    password = os.environ.get('SMARTBI_PASSWORD')
    base_url = os.environ.get('SMARTBI_BASE_URL', 'https://bi.61info.cn/smartbi/vision/index.jsp')

    if not username:
        username = input("请输入 Smartbi 用户名: ")
    if not password:
        password = getpass("请输入 Smartbi 密码: ")

    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)

    return {
        'username': username,
        'password': password,
        'base_url': base_url,
        'download_dir': DOWNLOAD_DIR,
        'screenshot_dir': SCREENSHOT_DIR
    }


async def export_single_report(page, config, report_url, business_name) -> Path:
    """导出单个报表，复用已登录的页面上下文"""
    download_file_path = None

    async def handle_download(download):
        nonlocal download_file_path
        file_name = download.suggested_filename or f"report_{int(time.time())}.xlsx"
        date_suffix = time.strftime("%y%m%d")
        name_parts = file_name.rsplit('.', 1)
        
        if len(name_parts) > 1:
            base_name = f"{name_parts[0]}_{date_suffix}"
            extension = name_parts[1]
        else:
            base_name = f"{file_name}_{date_suffix}"
            extension = "xlsx"
        
        download_file_path = config['download_dir'] / f"{base_name}.{extension}"
        counter = 1
        while download_file_path.exists():
            download_file_path = config['download_dir'] / f"{base_name}_{counter}.{extension}"
            counter += 1
        
        await download.save_as(str(download_file_path))
        print(f"[DEBUG] 文件下载完成: {download_file_path.name}")

    print(f"\n[INFO] 访问报表: {business_name}")
    await page.goto(report_url, wait_until="networkidle")

    try:
        await page.wait_for_selector('table', timeout=60000)
    except:
        pass

    await page.wait_for_load_state("networkidle", timeout=60000)
    print(f"[INFO] {business_name} 报表加载完成")

    print(f"[INFO] 导出 {business_name}...")

    page.on("download", handle_download)

    print("[DEBUG] 点击导出按钮")
    export_selectors = ['input:has-text("导出")', 'button:has-text("导出")', '[class*="export"]']
    export_found = False
    for selector in export_selectors:
        if await page.locator(selector).count() > 0:
            await page.click(selector)
            export_found = True
            break
    
    if not export_found:
        raise Exception("未找到导出按钮")

    print("[DEBUG] 选择 Excel 格式")
    try:
        await page.wait_for_selector('span:has-text("Excel")', timeout=5000)
        await page.click('span:has-text("Excel")')
    except PlaywrightTimeoutError:
        raise Exception("未找到 Excel 选项")

    print("[DEBUG] 等待在线导出按钮")
    online_export_selector = 'input[value="在线导出"]'
    
    try:
        await page.wait_for_selector(online_export_selector, timeout=10000)
        await page.click(online_export_selector)
        print("[DEBUG] 已点击在线导出")
    except PlaywrightTimeoutError:
        try:
            await page.wait_for_selector('input[atp="baseDialog_btnOK"]', timeout=5000)
            await page.click('input[atp="baseDialog_btnOK"]')
            print("[DEBUG] 已点击在线导出(备用选择器)")
        except PlaywrightTimeoutError:
            raise Exception("未找到在线导出按钮")

    print("[INFO] 等待下载完成...")
    download_timeout = 180
    start_time = time.time()
    
    while time.time() - start_time < download_timeout:
        if download_file_path and download_file_path.exists():
            break
        await asyncio.sleep(1)
        if int(time.time() - start_time) % 30 == 0:
            print(f"[INFO] 等待中... ({int(time.time() - start_time)}秒)")
    
    if not download_file_path or not download_file_path.exists():
        raise Exception(f"下载超时！{business_name} 报表未成功下载")

    print(f"[SUCCESS] {business_name} 导出完成: {download_file_path.name}")
    return download_file_path


async def login(page, config):
    """统一登录 - 只执行一次"""
    print("\n" + "="*60)
    print("统一登录 Smartbi (仅登录1次)")
    print("="*60)
    
    await page.goto(config['base_url'], wait_until="networkidle")

    try:
        await page.wait_for_selector('input[bofid="username"]', timeout=15000)
    except PlaywrightTimeoutError:
        raise Exception("登录页面加载超时")

    await page.fill('input[bofid="username"]', config['username'])
    await page.fill('input[bofid="password"]', config['password'])
    await page.click('input[bofid="login"]')

    await page.wait_for_load_state("networkidle", timeout=30000)
    print("[SUCCESS] 登录成功")


def analyze_data(file_path):
    wb = load_workbook(file_path, data_only=True)

    target_sheet = None
    for sheet_name in wb.sheetnames:
        if '小结指标' in sheet_name or '小结' in sheet_name:
            target_sheet = sheet_name
            break

    if not target_sheet:
        target_sheet = wb.sheetnames[0]

    ws = wb[target_sheet]
    all_data = list(ws.iter_rows(values_only=True))
    wb.close()
    return all_data


def parse_percent(value):
    if value is None:
        return None
    try:
        if isinstance(value, str):
            value = value.replace('%', '').replace(',', '').strip()
            return float(value) / 100
        return value
    except:
        return value


def extract_metrics(raw_data, extract_channels=True, channel_type='dual', channel_name='社群', kol_channel='KOL', offline_channel='线下'):
    data = {
        '主投ROI': None,
        '主投ROI环比': None,
        '约课数达成率': None,
        '约课数环比': None,
        '约课率达成率': None,
        '约课率环比': None,
        '约课成本环比': None,
        '约课成本Gap': None,
        '分发产能达成率': None,
        '分发产能环比': None,
    }

    if extract_channels:
        if channel_type == 'community':
            # 欧美澳商务：提取社群渠道
            data.update({
                '社群主投ROI': None,
                '社群主投ROI环比': None,
                '社群约课数达成率': None,
                '社群约课数环比': None,
                '社群约课率达成率': None,
                '社群约课率环比': None,
                '社群约课成本环比': None,
                '社群约课成本Gap': None,
                '社群分发产能达成率': None,
                '社群分发产能环比': None,
            })
        else:
            # 港澳商务：提取KOL和线下两个渠道
            data.update({
                'KOL主投ROI': None,
                'KOL主投ROI环比': None,
                'KOL约课数达成率': None,
                'KOL约课数环比': None,
                'KOL约课率达成率': None,
                'KOL约课率环比': None,
                'KOL约课成本环比': None,
                'KOL约课成本Gap': None,
                'KOL分发产能达成率': None,
                'KOL分发产能环比': None,
                '线下主投ROI': None,
                '线下主投ROI环比': None,
                '线下约课数达成率': None,
                '线下约课数环比': None,
                '线下约课率达成率': None,
                '线下约课率环比': None,
                '线下约课成本环比': None,
                '线下约课成本Gap': None,
                '线下分发产能达成率': None,
                '线下分发产能环比': None,
            })

    current_main_type = None
    current_channel = None

    for row_idx, row in enumerate(raw_data):
        if len(row) < 3:
            continue

        main_type = row[1]
        channel = row[2]
        supplier = row[3] if len(row) > 3 else None

        if main_type == '主投':
            continue

        if main_type is not None and main_type != '':
            current_main_type = main_type

        if channel is not None and channel != '':
            current_channel = channel

        if channel_type == 'community':
            # 欧美澳商务的提取逻辑（与单独脚本一致）
            if current_main_type == '思维' and channel == '总计':
                if len(row) > 3: data['主投ROI'] = parse_percent(row[3])
                if len(row) > 4: data['主投ROI环比'] = parse_percent(row[4])
                if len(row) > 5: data['约课数达成率'] = parse_percent(row[5])
                if len(row) > 6: data['约课数环比'] = parse_percent(row[6])
                if len(row) > 7: data['约课率达成率'] = parse_percent(row[7])
                if len(row) > 8: data['约课率环比'] = parse_percent(row[8])
                if len(row) > 11: data['约课成本环比'] = parse_percent(row[11])
                if len(row) > 12: data['约课成本Gap'] = row[12]
                if len(row) > 13: data['分发产能达成率'] = parse_percent(row[13])
                if len(row) > 14: data['分发产能环比'] = parse_percent(row[14])

            if extract_channels and current_main_type == '思维' and channel == channel_name:
                if len(row) > 3: data['社群主投ROI'] = parse_percent(row[3])
                if len(row) > 4: data['社群主投ROI环比'] = parse_percent(row[4])
                if len(row) > 5: data['社群约课数达成率'] = parse_percent(row[5])
                if len(row) > 6: data['社群约课数环比'] = parse_percent(row[6])
                if len(row) > 7: data['社群约课率达成率'] = parse_percent(row[7])
                if len(row) > 8: data['社群约课率环比'] = parse_percent(row[8])
                if len(row) > 11: data['社群约课成本环比'] = parse_percent(row[11])
                if len(row) > 12: data['社群约课成本Gap'] = row[12]
                if len(row) > 13: data['社群分发产能达成率'] = parse_percent(row[13])
                if len(row) > 14: data['社群分发产能环比'] = parse_percent(row[14])
        else:
            # 港澳商务的提取逻辑
            if current_main_type == '思维' and current_channel == '总计' and supplier == '总计':
                if len(row) > 4: data['主投ROI'] = parse_percent(row[4])
                if len(row) > 5: data['主投ROI环比'] = parse_percent(row[5])
                if len(row) > 6: data['约课数达成率'] = parse_percent(row[6])
                if len(row) > 7: data['约课数环比'] = parse_percent(row[7])
                if len(row) > 8: data['约课率达成率'] = parse_percent(row[8])
                if len(row) > 9: data['约课率环比'] = parse_percent(row[9])
                if len(row) > 12: data['约课成本环比'] = parse_percent(row[12])
                if len(row) > 13: data['约课成本Gap'] = row[13]
                if len(row) > 14: data['分发产能达成率'] = parse_percent(row[14])
                if len(row) > 15: data['分发产能环比'] = parse_percent(row[15])

            if extract_channels:
                if current_main_type == '思维' and current_channel == kol_channel and supplier == '总计':
                    if len(row) > 4: data['KOL主投ROI'] = parse_percent(row[4])
                    if len(row) > 5: data['KOL主投ROI环比'] = parse_percent(row[5])
                    if len(row) > 6: data['KOL约课数达成率'] = parse_percent(row[6])
                    if len(row) > 7: data['KOL约课数环比'] = parse_percent(row[7])
                    if len(row) > 8: data['KOL约课率达成率'] = parse_percent(row[8])
                    if len(row) > 9: data['KOL约课率环比'] = parse_percent(row[9])
                    if len(row) > 12: data['KOL约课成本环比'] = parse_percent(row[12])
                    if len(row) > 13: data['KOL约课成本Gap'] = row[13]
                    if len(row) > 14: data['KOL分发产能达成率'] = parse_percent(row[14])
                    if len(row) > 15: data['KOL分发产能环比'] = parse_percent(row[15])

                if current_main_type == '思维' and current_channel == offline_channel and supplier == '总计':
                    if len(row) > 4: data['线下主投ROI'] = parse_percent(row[4])
                    if len(row) > 5: data['线下主投ROI环比'] = parse_percent(row[5])
                    if len(row) > 6: data['线下约课数达成率'] = parse_percent(row[6])
                    if len(row) > 7: data['线下约课数环比'] = parse_percent(row[7])
                    if len(row) > 8: data['线下约课率达成率'] = parse_percent(row[8])
                    if len(row) > 9: data['线下约课率环比'] = parse_percent(row[9])
                    if len(row) > 12: data['线下约课成本环比'] = parse_percent(row[12])
                    if len(row) > 13: data['线下约课成本Gap'] = row[13]
                    if len(row) > 14: data['线下分发产能达成率'] = parse_percent(row[14])
                    if len(row) > 15: data['线下分发产能环比'] = parse_percent(row[15])

    return data


def get_color(value, metric_type='default'):
    if value is None:
        return 'black'
    try:
        if isinstance(value, str):
            value = value.replace('%', '').replace('+', '').strip()
        num_value = float(value)

        if metric_type == 'achievement':
            return 'green' if num_value >= 1 else 'red'
        elif metric_type == 'cost':
            return 'green' if num_value < 0 else 'red'
        elif metric_type == 'gap':
            if num_value > 0: return 'red'
            elif num_value < 0: return 'green'
            else: return 'black'
        else:
            return 'green' if num_value >= 0 else 'red'
    except:
        return 'black'


def format_value(value, is_gap=False):
    if value is None:
        return '-'
    try:
        if isinstance(value, str):
            return value
        return f"{round(value)}" if is_gap else f"{value:.2f}"
    except:
        return str(value)


def format_percent(value, show_sign=True):
    if value is None:
        return '-'
    try:
        if isinstance(value, str):
            return value
        formatted = f"{value*100:.2f}%"
        if show_sign and value > 0 and not formatted.startswith('+'):
            formatted = '+' + formatted
        return formatted
    except:
        return str(value)


def generate_summary(data, business_title, extract_channels=True, channel_type='dual'):
    summary = []
    summary.append({'type': 'title', 'content': business_title})

    overall_parts = []
    roi_val = format_value(data.get('主投ROI'))
    roi_change = format_percent(data.get('主投ROI环比'))
    roi_change_color = get_color(data.get('主投ROI环比'), 'default')
    overall_parts.append(f"主投ROI: {roi_val} (环比{{{roi_change_color}|{roi_change}}})，")

    enroll_achievement = format_percent(data.get('约课数达成率'), show_sign=False)
    enroll_achievement_color = get_color(data.get('约课数达成率'), 'achievement')
    enroll_change = format_percent(data.get('约课数环比'))
    enroll_change_color = get_color(data.get('约课数环比'), 'default')
    overall_parts.append(f"约课数 (达成{{{enroll_achievement_color}|{enroll_achievement}}}，环比{{{enroll_change_color}|{enroll_change}}})，")

    rate_achievement = format_percent(data.get('约课率达成率'), show_sign=False)
    rate_achievement_color = get_color(data.get('约课率达成率'), 'achievement')
    rate_change = format_percent(data.get('约课率环比'))
    rate_change_color = get_color(data.get('约课率环比'), 'default')
    overall_parts.append(f"约课率 (达成{{{rate_achievement_color}|{rate_achievement}}}，环比{{{rate_change_color}|{rate_change}}})，")

    cost_change = format_percent(data.get('约课成本环比'))
    cost_change_color = get_color(data.get('约课成本环比'), 'cost')
    cost_gap = format_value(data.get('约课成本Gap'), is_gap=True)
    cost_gap_color = get_color(data.get('约课成本Gap'), 'gap')
    overall_parts.append(f"约课成本 (环比{{{cost_change_color}|{cost_change}}}，gap{{{cost_gap_color}|{cost_gap}}})，")

    capacity_achievement = format_percent(data.get('分发产能达成率'), show_sign=False)
    capacity_achievement_color = get_color(data.get('分发产能达成率'), 'achievement')
    capacity_change = format_percent(data.get('分发产能环比'))
    capacity_change_color = get_color(data.get('分发产能环比'), 'default')
    overall_parts.append(f"分发产能 (达成{{{capacity_achievement_color}|{capacity_achievement}}}，环比{{{capacity_change_color}|{capacity_change}}})；")

    summary.append({'type': 'overall', 'content': ''.join(overall_parts)})

    if extract_channels:
        if channel_type == 'community':
            # 欧美澳商务：生成社群渠道小结
            community_parts = []
            comm_roi_val = format_value(data.get('社群主投ROI'))
            comm_roi_change = format_percent(data.get('社群主投ROI环比'))
            comm_roi_change_color = get_color(data.get('社群主投ROI环比'), 'default')
            community_parts.append(f"{{red|核心关注:}} 社群主投ROI {comm_roi_val} (环比{{{comm_roi_change_color}|{comm_roi_change}}})，")

            comm_enroll_achievement = format_percent(data.get('社群约课数达成率'), show_sign=False)
            comm_enroll_achievement_color = get_color(data.get('社群约课数达成率'), 'achievement')
            comm_enroll_change = format_percent(data.get('社群约课数环比'))
            comm_enroll_change_color = get_color(data.get('社群约课数环比'), 'default')
            community_parts.append(f"约课数 (达成{{{comm_enroll_achievement_color}|{comm_enroll_achievement}}}，环比{{{comm_enroll_change_color}|{comm_enroll_change}}})，")

            comm_rate_achievement = format_percent(data.get('社群约课率达成率'), show_sign=False)
            comm_rate_achievement_color = get_color(data.get('社群约课率达成率'), 'achievement')
            comm_rate_change = format_percent(data.get('社群约课率环比'))
            comm_rate_change_color = get_color(data.get('社群约课率环比'), 'default')
            community_parts.append(f"约课率 (达成{{{comm_rate_achievement_color}|{comm_rate_achievement}}}，环比{{{comm_rate_change_color}|{comm_rate_change}}})，")

            comm_cost_change = format_percent(data.get('社群约课成本环比'))
            comm_cost_change_color = get_color(data.get('社群约课成本环比'), 'cost')
            comm_cost_gap = format_value(data.get('社群约课成本Gap'), is_gap=True)
            comm_cost_gap_color = get_color(data.get('社群约课成本Gap'), 'gap')
            community_parts.append(f"约课成本 (环比{{{comm_cost_change_color}|{comm_cost_change}}}，gap{{{comm_cost_gap_color}|{comm_cost_gap}}})，")

            comm_capacity_achievement = format_percent(data.get('社群分发产能达成率'), show_sign=False)
            comm_capacity_achievement_color = get_color(data.get('社群分发产能达成率'), 'achievement')
            comm_capacity_change = format_percent(data.get('社群分发产能环比'))
            comm_capacity_change_color = get_color(data.get('社群分发产能环比'), 'default')
            community_parts.append(f"分发产能 (达成{{{comm_capacity_achievement_color}|{comm_capacity_achievement}}}，环比{{{comm_capacity_change_color}|{comm_capacity_change}}})；")

            summary.append({'type': 'community', 'content': ''.join(community_parts)})
        else:
            # 港澳商务：生成KOL和线下渠道小结
            kol_parts = []
            kol_roi_val = format_value(data.get('KOL主投ROI'))
            kol_roi_change = format_percent(data.get('KOL主投ROI环比'))
            kol_roi_change_color = get_color(data.get('KOL主投ROI环比'), 'default')
            kol_parts.append(f"{{red|核心关注:}} KOL主投ROI {kol_roi_val} (环比{{{kol_roi_change_color}|{kol_roi_change}}})，")

            kol_enroll_achievement = format_percent(data.get('KOL约课数达成率'), show_sign=False)
            kol_enroll_achievement_color = get_color(data.get('KOL约课数达成率'), 'achievement')
            kol_enroll_change = format_percent(data.get('KOL约课数环比'))
            kol_enroll_change_color = get_color(data.get('KOL约课数环比'), 'default')
            kol_parts.append(f"约课数 (达成{{{kol_enroll_achievement_color}|{kol_enroll_achievement}}}，环比{{{kol_enroll_change_color}|{kol_enroll_change}}})，")

            kol_rate_achievement = format_percent(data.get('KOL约课率达成率'), show_sign=False)
            kol_rate_achievement_color = get_color(data.get('KOL约课率达成率'), 'achievement')
            kol_rate_change = format_percent(data.get('KOL约课率环比'))
            kol_rate_change_color = get_color(data.get('KOL约课率环比'), 'default')
            kol_parts.append(f"约课率 (达成{{{kol_rate_achievement_color}|{kol_rate_achievement}}}，环比{{{kol_rate_change_color}|{kol_rate_change}}})，")

            kol_cost_change = format_percent(data.get('KOL约课成本环比'))
            kol_cost_change_color = get_color(data.get('KOL约课成本环比'), 'cost')
            kol_cost_gap = format_value(data.get('KOL约课成本Gap'), is_gap=True)
            kol_cost_gap_color = get_color(data.get('KOL约课成本Gap'), 'gap')
            kol_parts.append(f"约课成本 (环比{{{kol_cost_change_color}|{kol_cost_change}}}，gap{{{kol_cost_gap_color}|{kol_cost_gap}}})，")

            kol_capacity_achievement = format_percent(data.get('KOL分发产能达成率'), show_sign=False)
            kol_capacity_achievement_color = get_color(data.get('KOL分发产能达成率'), 'achievement')
            kol_capacity_change = format_percent(data.get('KOL分发产能环比'))
            kol_capacity_change_color = get_color(data.get('KOL分发产能环比'), 'default')
            kol_parts.append(f"分发产能 (达成{{{kol_capacity_achievement_color}|{kol_capacity_achievement}}}，环比{{{kol_capacity_change_color}|{kol_capacity_change}}})；")

            summary.append({'type': 'community', 'content': ''.join(kol_parts)})

            offline_parts = []
            offline_roi_val = format_value(data.get('线下主投ROI'))
            offline_roi_change = format_percent(data.get('线下主投ROI环比'))
            offline_roi_change_color = get_color(data.get('线下主投ROI环比'), 'default')
            offline_parts.append(f"{' '*6}线下主投ROI {offline_roi_val} (环比{{{offline_roi_change_color}|{offline_roi_change}}})，")

            offline_enroll_achievement = format_percent(data.get('线下约课数达成率'), show_sign=False)
            offline_enroll_achievement_color = get_color(data.get('线下约课数达成率'), 'achievement')
            offline_enroll_change = format_percent(data.get('线下约课数环比'))
            offline_enroll_change_color = get_color(data.get('线下约课数环比'), 'default')
            offline_parts.append(f"约课数 (达成{{{offline_enroll_achievement_color}|{offline_enroll_achievement}}}，环比{{{offline_enroll_change_color}|{offline_enroll_change}}})，")

            offline_rate_achievement = format_percent(data.get('线下约课率达成率'), show_sign=False)
            offline_rate_achievement_color = get_color(data.get('线下约课率达成率'), 'achievement')
            offline_rate_change = format_percent(data.get('线下约课率环比'))
            offline_rate_change_color = get_color(data.get('线下约课率环比'), 'default')
            offline_parts.append(f"约课率 (达成{{{offline_rate_achievement_color}|{offline_rate_achievement}}}，环比{{{offline_rate_change_color}|{offline_rate_change}}})，")

            offline_cost_change = format_percent(data.get('线下约课成本环比'))
            offline_cost_change_color = get_color(data.get('线下约课成本环比'), 'cost')
            offline_cost_gap = format_value(data.get('线下约课成本Gap'), is_gap=True)
            offline_cost_gap_color = get_color(data.get('线下约课成本Gap'), 'gap')
            offline_parts.append(f"约课成本 (环比{{{offline_cost_change_color}|{offline_cost_change}}}，gap{{{offline_cost_gap_color}|{offline_cost_gap}}})，")

            offline_capacity_achievement = format_percent(data.get('线下分发产能达成率'), show_sign=False)
            offline_capacity_achievement_color = get_color(data.get('线下分发产能达成率'), 'achievement')
            offline_capacity_change = format_percent(data.get('线下分发产能环比'))
            offline_capacity_change_color = get_color(data.get('线下分发产能环比'), 'default')
            offline_parts.append(f"分发产能 (达成{{{offline_capacity_achievement_color}|{offline_capacity_achievement}}}，环比{{{offline_capacity_change_color}|{offline_capacity_change}}})；")

            summary.append({'type': 'community', 'content': ''.join(offline_parts)})

    return summary


def save_summary_to_excel(summaries):
    date_suffix = time.strftime("%y%m%d")
    output_file = OUTPUT_DIR / f"商务周报数据小结_{date_suffix}.xlsx"

    title_font = Font(bold=True, size=10, color="000000", name=FEISHU_DEFAULT_FONT)
    normal_inline_font = InlineFont(rFont=FEISHU_DEFAULT_FONT, sz=10, color="000000")
    green_inline_font = InlineFont(rFont=FEISHU_DEFAULT_FONT, sz=10, color=GREEN_COLOR)
    red_inline_font = InlineFont(rFont=FEISHU_DEFAULT_FONT, sz=10, color=RED_COLOR)

    white_border = Border(
        left=Side(style='thin', color='FFFFFF'),
        right=Side(style='thin', color='FFFFFF'),
        top=Side(style='thin', color='FFFFFF'),
        bottom=Side(style='thin', color='FFFFFF')
    )

    def parse_colored_text(text):
        parts = []
        i = 0
        while i < len(text):
            if text[i] == '{':
                end_idx = text.find('}', i)
                if end_idx != -1:
                    if i > 0:
                        parts.append({'text': text[:i], 'font': normal_inline_font})
                    content = text[i+1:end_idx]
                    pipe_idx = content.find('|')
                    if pipe_idx != -1:
                        color_name = content[:pipe_idx]
                        colored_text = content[pipe_idx+1:]
                        if color_name == 'green':
                            parts.append({'text': colored_text, 'font': green_inline_font})
                        elif color_name == 'red':
                            parts.append({'text': colored_text, 'font': red_inline_font})
                        else:
                            parts.append({'text': colored_text, 'font': normal_inline_font})
                    text = text[end_idx+1:]
                    i = 0
                    continue
            i += 1
        if text:
            parts.append({'text': text, 'font': normal_inline_font})
        return parts

    def write_summary_to_sheet(ws, summary):
        row = 1
        for item in summary:
            cell = ws.cell(row=row, column=1)
            cell.alignment = Alignment(wrap_text=False)
            cell.border = white_border

            if item['type'] == 'title':
                cell.font = title_font
                cell.value = item['content']
            else:
                parts = parse_colored_text(item['content'])
                if len(parts) == 1:
                    cell.font = Font(size=10, name=FEISHU_DEFAULT_FONT, color=parts[0]['font'].color)
                    cell.value = parts[0]['text']
                else:
                    rich_parts = [TextBlock(part['font'], part['text']) for part in parts]
                    cell.value = CellRichText(rich_parts)
                    # Rich text run color/font is controlled by InlineFont.
                    # Avoid overriding with cell.font, otherwise all runs may become one style.
            row += 1
        ws.column_dimensions['A'].width = 150

    all_summaries = []
    for business_name, summary in summaries.items():
        config = BUSINESS_CONFIGS[business_name]
        all_summaries.append((business_name, config['title'], summary, config['extract_channels']))

    if output_file.exists():
        wb = load_workbook(output_file)
    else:
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

    generated_sheets = []
    for business_name, title, summary, extract_channels in all_summaries:
        sheet_name = business_name
        if sheet_name in wb.sheetnames:
            existing_versions = []
            for name in wb.sheetnames:
                if name.startswith(f"{sheet_name}v"):
                    try:
                        ver = int(name.replace(f"{sheet_name}v", ''))
                        existing_versions.append(ver)
                    except:
                        pass
            new_version = max(existing_versions) + 1 if existing_versions else 1
            sheet_name = f"{sheet_name}v{new_version}"

        ws = wb.create_sheet(title=sheet_name)
        write_summary_to_sheet(ws, summary)
        generated_sheets.append(sheet_name)

    wb.save(output_file)
    return output_file, generated_sheets


async def main(selected_businesses=None):
    print("=" * 70)
    print("商务周报数据小结一键生成脚本")
    print("=" * 70)

    if selected_businesses:
        businesses_to_run = {k: v for k, v in BUSINESS_CONFIGS.items() if k in selected_businesses}
    else:
        businesses_to_run = BUSINESS_CONFIGS

    print("\n将依次处理以下商务:")
    for name in businesses_to_run.keys():
        print(f"  - {name}")
    print()

    config = load_config()

    playwright = await async_playwright().start()
    browser = await playwright.chromium.launch(
        headless=False,
        args=["--start-maximized", "--disable-popup-blocking"]
    )
    context = await browser.new_context(viewport={"width": 1920, "height": 1080})
    page = await context.new_page()
    
    page.on("dialog", lambda dialog: dialog.accept())

    try:
        await login(page, config)

        all_summaries = {}

        for business_name, business_config in businesses_to_run.items():
            try:
                download_path = await export_single_report(
                    page,
                    config,
                    business_config['report_url'],
                    business_name
                )

                raw_data = analyze_data(download_path)
                metrics = extract_metrics(
                    raw_data,
                    business_config['extract_channels'],
                    business_config.get('channel_type', 'dual'),
                    business_config.get('channel_name', '社群'),
                    business_config.get('kol_channel', 'KOL'),
                    business_config.get('offline_channel', '线下')
                )
                summary = generate_summary(
                    metrics,
                    business_config['title'],
                    business_config['extract_channels'],
                    business_config.get('channel_type', 'dual')
                )

                all_summaries[business_name] = summary
                print(f"\n[INFO] {business_name} 小结生成完成")

            except Exception as e:
                print(f"[ERROR] {business_name} 处理失败: {e}")
                traceback.print_exc()
                all_summaries[business_name] = None
    finally:
        await context.close()
        await browser.close()
        await playwright.stop()

    print(f"\n{'='*60}")
    print("保存所有小结到Excel...")
    print(f"{'='*60}")

    valid_summaries = {k: v for k, v in all_summaries.items() if v is not None}
    if valid_summaries:
        output_file, generated_sheets = save_summary_to_excel(valid_summaries)
        print(f"\n[SUCCESS] 所有小结已保存到: {output_file}")
        
        print("\n✅ 本次生成的Sheet明细:")
        for sheet_name in generated_sheets:
            print(f"  - {sheet_name}")
    else:
        print("\n[WARNING] 没有成功生成任何小结")

    print("\n处理状态汇总:")
    for name in businesses_to_run.keys():
        status = "✓ 成功" if all_summaries.get(name) else "✗ 失败"
        print(f"  - {name}: {status}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='商务周报数据小结生成脚本')
    parser.add_argument('--businesses', '-b', nargs='+',
                        choices=['欧美澳商务', '港澳商务', 'Local商务'],
                        help='指定要处理的商务类型，默认处理全部')
    args = parser.parse_args()

    asyncio.run(main(args.businesses))
