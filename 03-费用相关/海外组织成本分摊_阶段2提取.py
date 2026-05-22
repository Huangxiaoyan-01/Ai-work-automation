import re
import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import numbers

DOWNLOAD_DIR = Path(r"C:\Users\huangxiaoyan\Documents\trae_projects\test_week\downloads")
OUTPUT_DIR = Path(r"C:\Users\huangxiaoyan\Documents\trae_projects\test_week\output")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

REGIONS = ["整体", "非台湾", "港澳", "台湾"]
REPORT_NAME = "海外漏斗达成情况-月趋势"

# 上表口径：展示名称 -> 在导出文件中匹配的渠道关键词
TOP_CHANNEL_MAP = {
    "投放主投": ["投放主投"],
    "海外商务主投": ["海外商务汇总"],
    "港澳商务主投": ["海外港澳商务主投"],
    "台湾商务主投": ["海外台湾商务主投"],
}

# 下表口径（非台湾：思维/美术）
BOTTOM_CHANNEL_MAP = {
    "转联投总计": ["转介绍联投汇总"],
    "投放总计": ["投放汇总", "投放总计"],
    "欧美澳商务总计": ["海外商务汇总"],
    "港澳商务总计": ["港澳商务汇总"],
}


def normalize(v: str) -> str:
    return re.sub(r"\s+", "", str(v or ""))


def get_default_month() -> str:
    today = datetime.date.today()
    y, m = today.year, today.month
    if m == 1:
        return f"{y-1}-12"
    return f"{y}-{m-1:02d}"


def extract_version_num(path: Path) -> int:
    m = re.search(r"v(\d+)\.xlsx$", path.name)
    return int(m.group(1)) if m else 0


def pick_latest_report(region: str) -> Path | None:
    pattern = f"{REPORT_NAME}_{region}_*.xlsx"
    files = list(DOWNLOAD_DIR.glob(pattern))
    if not files:
        return None

    def sort_key(p: Path):
        # 名称里默认包含 yymmdd，优先按日期+版本排序
        m = re.search(r"_(\d{6})(?:v(\d+))?\.xlsx$", p.name)
        if m:
            date_num = int(m.group(1))
            ver = int(m.group(2)) if m.group(2) else 0
            return (date_num, ver)
        return (0, extract_version_num(p))

    return sorted(files, key=sort_key)[-1]


def find_target_col(ws, month_str: str) -> int | None:
    candidates = {month_str, month_str.replace("-0", "-"), month_str.replace("-", "/")}

    # 优先定位“例子数”表头行（B列=例子数，D列=指标）
    # 你提供的文件里该表头通常不在前15行，而在60行附近。
    header_row = None
    scan_rows = min(ws.max_row, 200)
    for row in range(1, scan_rows + 1):
        b_val = normalize(ws.cell(row=row, column=2).value)
        d_val = normalize(ws.cell(row=row, column=4).value)
        if b_val == normalize("例子数") and d_val == normalize("指标"):
            header_row = row
            break

    # 若定位到“例子数”表头，则只在该行找目标月份列
    if header_row:
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row, column=col).value
            if val is None:
                continue
            txt = str(val).strip()
            if any(c in txt for c in candidates):
                return col

    # 兜底：全表前200行扫描月份列
    for row in range(1, scan_rows + 1):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            txt = str(val).strip()
            if any(c in txt for c in candidates):
                return col
    return None


def extract_sheet_channel_values(file_path: Path, month_str: str) -> dict[str, list[dict]]:
    wb = load_workbook(file_path, data_only=True)
    result: dict[str, list[dict]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 美术sheet可能隐藏了D列（指标列），先取消隐藏再取值
        try:
            if ws.column_dimensions["D"].hidden:
                ws.column_dimensions["D"].hidden = False
                print(f"[DEBUG] 已取消隐藏列D: {sheet_name}")
        except Exception:
            pass

        col = find_target_col(ws, month_str)
        result[sheet_name] = []
        if not col:
            continue

        # B列在Excel里存在大量合并单元格，明细行会是空值。
        # 这里维护“当前分组”，保证投放/转介绍/海外商务等分组信息不丢失。
        current_group = ""

        for row in range(1, ws.max_row + 1):
            # C列是渠道明细（如“投放汇总”），B列是上级分组，优先用C列
            channel_c = ws.cell(row=row, column=3).value
            channel_b = ws.cell(row=row, column=2).value
            if channel_b not in (None, ""):
                current_group = str(channel_b).strip()
            indicator = ws.cell(row=row, column=4).value
            # 兼容隐藏空格/前后缀文本，按“包含例子数”识别
            if "例子数" not in normalize(indicator):
                continue
            value = ws.cell(row=row, column=col).value
            result[sheet_name].append({
                "group": current_group,
                "channel": str(channel_c or channel_b or "").strip(),
                "value": float(value) if value not in (None, "") else 0.0,
            })
    return result


def sum_by_keywords(rows: list[dict], keywords: list[str]) -> float:
    total = 0.0
    for r in rows:
        c = normalize(r["channel"])
        if any(normalize(k) in c for k in keywords):
            total += float(r["value"] or 0)
    return total


def get_region_total(data_map: dict, region: str, keywords: list[str], sheet_name: str | None = None) -> float:
    if region not in data_map:
        return 0.0
    total = 0.0
    for _sheet, rows in data_map[region].items():
        if sheet_name and _sheet != sheet_name:
            continue
        total += sum_by_keywords(rows, keywords)
    return total


def get_region_invest_main_total(data_map: dict, region: str, sheet_name: str = "思维") -> float:
    """上表“投放主投”口径：优先取 C列=投放主投(汇总行)，无此行再回退到主投明细汇总。"""
    if region not in data_map:
        return 0.0
    rows = data_map[region].get(sheet_name, [])

    # 口径优先：直接取“投放主投”汇总行
    exact_total = 0.0
    for r in rows:
        channel_text = normalize(r.get("channel", ""))
        if channel_text in (normalize("投放主投"), normalize("投放主投汇总")):
            exact_total += float(r.get("value") or 0)
    if exact_total > 0:
        return exact_total

    # 回退：汇总投放分组下主投明细（不含辅投）
    total = 0.0
    for r in rows:
        group_text = normalize(r.get("group", ""))
        channel_text = normalize(r.get("channel", ""))
        if "投放" not in group_text:
            continue
        if "主投" in channel_text and "辅投" not in channel_text:
            total += float(r.get("value") or 0)
    return total


def get_non_tw_sheet_total(data_map: dict, sheet_name: str, keywords: list[str]) -> float:
    non_tw = data_map.get("非台湾", {})
    # 严格优先：按sheet名精确取（需求指定C列必须取美术sheet）
    rows = non_tw.get(sheet_name, [])

    # 次优：sheet名归一化匹配（防止名称有空格）
    if not rows:
        alias = normalize(sheet_name)
        for k, v in non_tw.items():
            if normalize(k) == alias:
                rows = v
                break

    # 最后兜底：仅在请求“美术”且确实找不到时，取第二个sheet并打印告警
    if not rows and non_tw and normalize(sheet_name) == normalize("美术"):
        sheet_keys = list(non_tw.keys())
        if len(sheet_keys) >= 2:
            rows = non_tw.get(sheet_keys[1], [])
            print(f"[WARN] 未精确命中美术sheet，兜底使用: {sheet_keys[1]}")
    return sum_by_keywords(rows, keywords)


def get_non_tw_metric_total(data_map: dict, sheet_name: str, metric_name: str) -> float:
    """下表口径：仅用于美术sheet的精确行匹配（B/C列精确值）。"""
    non_tw = data_map.get("非台湾", {})
    rows = non_tw.get(sheet_name, [])
    if not rows:
        alias = normalize(sheet_name)
        for k, v in non_tw.items():
            if normalize(k) == alias:
                rows = v
                break
    if not rows:
        return 0.0

    metric = normalize(metric_name)
    total = 0.0
    for r in rows:
        group_text = normalize(r.get("group", ""))
        channel_text = normalize(r.get("channel", ""))
        value = float(r.get("value") or 0)

        if metric == normalize("转联投总计"):
            # B=转介绍联投, C=汇总
            if group_text == normalize("转介绍联投") and channel_text == normalize("汇总"):
                total += value
        elif metric == normalize("投放总计"):
            # B=投放, C=汇总
            if group_text == normalize("投放") and channel_text == normalize("汇总"):
                total += value
        elif metric == normalize("欧美澳商务总计"):
            # B=海外商务, C=汇总
            if group_text == normalize("海外商务") and channel_text == normalize("汇总"):
                total += value
        elif metric == normalize("港澳商务总计"):
            # B=海外港澳商务, C=汇总
            if group_text == normalize("海外港澳商务") and channel_text == normalize("汇总"):
                total += value
    return total


def prepare_output_wb() -> tuple:
    date_tag = datetime.date.today().strftime("%y%m%d")
    output_path = OUTPUT_DIR / f"海外组织成本分摊_{date_tag}.xlsx"

    if output_path.exists():
        wb = load_workbook(output_path)
        idx = 1
        while f"v{idx}" in wb.sheetnames:
            idx += 1
        ws = wb.create_sheet(f"v{idx}")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "v1"

    return wb, ws, output_path


def write_output(ws, data_map: dict, month_str: str) -> None:
    ws["A1"] = f"月份: {month_str}"

    # 上表
    ws["A3"] = ""
    ws["B3"] = "整体"
    ws["C3"] = "港澳"
    ws["D3"] = "台湾"
    ws["E3"] = "欧美澳"

    row = 4
    for display, keys in TOP_CHANNEL_MAP.items():
        ws.cell(row=row, column=1, value=display)
        if display == "投放主投":
            ws.cell(row=row, column=2, value=get_region_invest_main_total(data_map, "整体", sheet_name="思维"))
            ws.cell(row=row, column=3, value=get_region_invest_main_total(data_map, "港澳", sheet_name="思维"))
            ws.cell(row=row, column=4, value=get_region_invest_main_total(data_map, "台湾", sheet_name="思维"))
        else:
            ws.cell(row=row, column=2, value=get_region_total(data_map, "整体", keys, sheet_name="思维"))
            ws.cell(row=row, column=3, value=get_region_total(data_map, "港澳", keys, sheet_name="思维"))
            ws.cell(row=row, column=4, value=get_region_total(data_map, "台湾", keys, sheet_name="思维"))
        ws.cell(row=row, column=5, value=f"=B{row}-C{row}-D{row}")
        row += 1

    ws["A8"] = ""
    ws["B8"] = "整体"
    ws["C8"] = "港澳"
    ws["D8"] = "台湾"
    ws["E8"] = "欧美澳"

    for r in range(9, 13):
        src = r - 5
        ws.cell(row=r, column=1, value=ws.cell(row=src, column=1).value)
        ws.cell(row=r, column=2, value=f"=B{src}")
        ws.cell(row=r, column=3, value=f"=C{src}/$B{src}")
        ws.cell(row=r, column=4, value=f"=D{src}/$B{src}")
        ws.cell(row=r, column=5, value=f"=E{src}/$B{src}")
        ws.cell(row=r, column=3).number_format = "0%"
        ws.cell(row=r, column=4).number_format = "0%"
        ws.cell(row=r, column=5).number_format = "0%"

    # 下表
    ws["A14"] = "非台湾"
    ws["B14"] = "思维"
    ws["C14"] = "美术"

    row = 15
    for display, keys in BOTTOM_CHANNEL_MAP.items():
        ws.cell(row=row, column=1, value=display)
        # 思维列：保留原口径（关键词汇总）
        ws.cell(row=row, column=2, value=get_non_tw_sheet_total(data_map, "思维", keys))
        # 美术列：使用精确行口径，避免重复计算
        ws.cell(row=row, column=3, value=get_non_tw_metric_total(data_map, "美术", display))
        row += 1

    ws["B19"] = "=SUM(B15:B18)"
    ws["C19"] = "=SUM(C15:C18)"
    ws["D19"] = "=B19+C19"
    ws["B20"] = "=B19/$D$19"
    ws["B20"].number_format = "0%"


def main() -> None:
    month_input = input(f"请输入目标月份(YYYY-MM，默认上月 {get_default_month()}): ").strip()
    month_str = month_input or get_default_month()

    selected_files: dict[str, Path] = {}
    for region in REGIONS:
        p = pick_latest_report(region)
        if p:
            selected_files[region] = p

    if not selected_files:
        print("[ERROR] downloads 目录未找到可用报表，请先执行导出。")
        return

    print("\n[INFO] 本次使用文件:")
    for region, p in selected_files.items():
        print(f"  - {region}: {p.name}")

    data_map = {}
    for region, file_path in selected_files.items():
        data_map[region] = extract_sheet_channel_values(file_path, month_str)

    wb, ws, output_path = prepare_output_wb()
    write_output(ws, data_map, month_str)
    try:
        wb.save(output_path)
        final_path = output_path
    except PermissionError:
        ts = datetime.datetime.now().strftime("%H%M%S")
        fallback = OUTPUT_DIR / f"海外组织成本分摊_{datetime.date.today().strftime('%y%m%d')}_{ts}.xlsx"
        wb.save(fallback)
        final_path = fallback
        print(f"\n[WARN] 原文件被占用，已另存为: {fallback}")

    print(f"\n[SUCCESS] 输出完成: {final_path}")


if __name__ == "__main__":
    main()
